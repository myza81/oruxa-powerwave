"""Tests for cleaned data export (CSV/Excel ingestion Slice 12, the
2026-09-04 "export the resolved Time Axis" enhancement (DEC-074), and
the 2026-09-04 "manifest/provenance is optional" enhancement). Pure
service-level tests -- no HTTP; API-level coverage (response shape,
headers, HTTP status codes) lives in tests/test_preparation_sources_api.py's
own Slice 12 test classes.

Every pre-existing test in this file calls `export_preparation_source()`
with an explicit `mode=EXPORT_MODE_WITH_PROVENANCE` -- they test ZIP/
manifest contents, so they keep testing exactly that mode unchanged.
`TestDataOnlyExport` below covers the NEW default (`EXPORT_MODE_DATA_ONLY`)
mode; `TestModeEquivalence` confirms both modes agree on the cleaned
data itself.
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import zipfile

import pytest
from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from starlette.datastructures import Headers

from app.services.errors import (
    ExportNotReadyError,
    ExportRequiresIntervalError,
    SourceNotFoundError,
    WorksheetNotSelectedError,
)
from app.services.preparation_export_service import (
    EXPORT_MODE_DATA_ONLY,
    EXPORT_MODE_WITH_PROVENANCE,
    export_preparation_source,
)
from app.services.preparation_import_service import (
    import_csv_preparation_source,
    import_excel_preparation_source,
    select_preparation_worksheet,
)
from app.domain.working_overlay import column_key
from app.services.preparation_session_registry import PreparationSessionRegistry
from app.services.time_axis_service import interpret_time_axis, set_time_axis_configuration
from app.services.working_overlay_service import (
    edit_cell,
    set_column_engineering_quantity,
    set_column_measured_unit,
    set_column_role,
    set_data_region,
    set_header_row,
    set_row_excluded,
)

WS = "ws-1"


def _upload(content: bytes, filename: str, content_type: str) -> UploadFile:
    return UploadFile(file=io.BytesIO(content), filename=filename, headers=Headers({"content-type": content_type}))


def _add_csv(registry: PreparationSessionRegistry, content: bytes, filename: str = "e.csv") -> str:
    summary = asyncio.run(
        import_csv_preparation_source(
            workspace_id=WS, csv_upload=_upload(content, filename, "text/csv"),
            max_total_bytes=100 * 1024 * 1024, registry=registry,
        )
    )
    return summary.source_id


def _build_xlsx(sheets: dict) -> bytes:
    workbook = Workbook()
    names = list(sheets.keys())
    workbook.active.title = names[0]
    for row in sheets[names[0]]:
        workbook.active.append(row)
    for name in names[1:]:
        ws = workbook.create_sheet(name)
        for row in sheets[name]:
            ws.append(row)
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _add_excel(registry: PreparationSessionRegistry, content: bytes, filename: str = "e.xlsx") -> str:
    summary = asyncio.run(
        import_excel_preparation_source(
            workspace_id=WS,
            excel_upload=_upload(content, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            max_total_bytes=100 * 1024 * 1024, registry=registry,
        )
    )
    return summary.source_id


def _unzip(content: bytes) -> zipfile.ZipFile:
    return zipfile.ZipFile(io.BytesIO(content))


def _read_csv_rows(zf: zipfile.ZipFile) -> list[list[str]]:
    name = next(n for n in zf.namelist() if n.endswith(".csv"))
    text = zf.read(name).decode("utf-8")
    return list(csv.reader(io.StringIO(text)))


def _read_manifest(zf: zipfile.ZipFile) -> dict:
    name = next(n for n in zf.namelist() if n.endswith(".manifest.json"))
    return json.loads(zf.read(name))


def _read_xlsx_rows(zf: zipfile.ZipFile) -> list[tuple]:
    name = next(n for n in zf.namelist() if n.endswith(".xlsx"))
    wb = load_workbook(io.BytesIO(zf.read(name)))
    return list(wb.active.iter_rows(values_only=True))


def _mark_time_axis(registry: PreparationSessionRegistry, source_id: str, *column_indices: int) -> None:
    for column_index in column_indices:
        set_column_role(workspace_id=WS, source_id=source_id, column_index=column_index, role="time_axis", registry=registry)


def _mark_waveform(registry: PreparationSessionRegistry, source_id: str, *column_indices: int) -> None:
    for column_index in column_indices:
        set_column_role(workspace_id=WS, source_id=source_id, column_index=column_index, role="waveform", registry=registry)


def _confirm_absolute(registry, source_id, *, column_index=0, date_order=None):
    options = {"date_order": date_order} if date_order else {}
    return set_time_axis_configuration(
        workspace_id=WS, source_id=source_id, column_indices=(column_index,),
        interpreter_id="absolute_datetime", options=options, confirmed=True, registry=registry,
    )


def _confirm_split_date_time(registry, source_id, *, date_column, time_column, date_order):
    return set_time_axis_configuration(
        workspace_id=WS, source_id=source_id, column_indices=(date_column, time_column),
        interpreter_id="split_date_time", options={"date_order": date_order}, confirmed=True, registry=registry,
    )


def _confirm_elapsed(registry, source_id, *, column_index=0, unit="seconds"):
    return set_time_axis_configuration(
        workspace_id=WS, source_id=source_id, column_indices=(column_index,),
        interpreter_id="elapsed_numeric", unit=unit, confirmed=True, registry=registry,
    )


def _confirm_sample_index(registry, source_id, *, column_index=0, interval_seconds=None, confirmed=True):
    return set_time_axis_configuration(
        workspace_id=WS, source_id=source_id, column_indices=(column_index,),
        interpreter_id="sample_index", interval_seconds=interval_seconds, confirmed=confirmed, registry=registry,
    )


def _ready_absolute_source(registry, content: bytes, *, time_col=0, waveform_cols=(1,)) -> str:
    """A minimal, immediately-exportable CSV: an unambiguous ISO-style
    absolute Time Axis column, plus the given Waveform column(s)."""
    sid = _add_csv(registry, content)
    _mark_time_axis(registry, sid, time_col)
    _mark_waveform(registry, sid, *waveform_cols)
    _confirm_absolute(registry, sid, column_index=time_col)
    return sid


class TestExportGating:
    def test_unconfigured_time_axis_blocks_export(self):
        prep = PreparationSessionRegistry()
        sid = _add_csv(prep, b"1,2\n")
        _mark_waveform(prep, sid, 1)

        with pytest.raises(ExportNotReadyError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

    def test_missing_waveform_column_blocks_export(self):
        prep = PreparationSessionRegistry()
        sid = _add_csv(prep, b"2026-08-31 13:00:00,1\n")
        _mark_time_axis(prep, sid, 0)
        _confirm_absolute(prep, sid, column_index=0)

        with pytest.raises(ExportNotReadyError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

    # Preparation Status integrity guardrail: readiness now blocks a
    # `manual` configuration unconditionally (ISSUE_TIME_AXIS_MANUAL_
    # UNRESOLVED, app.services.readiness_service), so `_ensure_exportable()`'s
    # own `is_ready` check (checked first) now raises the generic
    # ExportNotReadyError before ever reaching the specific
    # ExportUnsupportedInterpreterError branch below it -- that branch is
    # retained as harmless defense-in-depth (see DECISIONS.md) but is no
    # longer reachable for `manual`/`unsupported` specifically.
    def test_manual_interpreter_blocks_export(self):
        prep = PreparationSessionRegistry()
        sid = _add_csv(prep, b"2026-08-31 13:00:00,1\n")
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        set_time_axis_configuration(
            workspace_id=WS, source_id=sid, column_indices=(0,), interpreter_id="manual",
            family="absolute", provenance="native", confirmed=True, registry=prep,
        )

        with pytest.raises(ExportNotReadyError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

    def test_sample_index_without_interval_blocks_export(self):
        prep = PreparationSessionRegistry()
        sid = _add_csv(prep, b"100,1.0\n101,2.0\n")
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_sample_index(prep, sid, interval_seconds=None)

        with pytest.raises(ExportRequiresIntervalError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

    def test_unconfirmed_reconstruction_blocks_export(self):
        prep = PreparationSessionRegistry()
        lines = [f"2026-08-31 13:00:{i // 2:02d},{i}.0" for i in range(6)]
        sid = _add_csv(prep, ("\n".join(lines) + "\n").encode())
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        set_time_axis_configuration(
            workspace_id=WS, source_id=sid, column_indices=(0,),
            interpreter_id="repeated_timestamp_precision_loss", confirmed=False, registry=prep,
        )

        with pytest.raises(ExportNotReadyError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

    def test_ready_absolute_source_exports_successfully(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1.0\n")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        assert result.filename.endswith(".zip")


class TestAbsoluteTimeExport:
    def test_ambiguous_date_order_resolved_exports_iso_timestamps(self):
        # Task section AC's own worked example.
        prep = PreparationSessionRegistry()
        content = b"Date,Time,Voltage\n3/6/26,18:04:00.000,132.1\n3/6/26,18:04:00.020,132.2\n3/6/26,18:04:00.040,132.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=4, registry=prep)
        _mark_time_axis(prep, sid, 0, 1)
        _mark_waveform(prep, sid, 2)
        _confirm_split_date_time(prep, sid, date_column=0, time_column=1, date_order="dmy")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert rows[0] == ["Time", "Voltage"]
        assert [r[0] for r in rows[1:]] == [
            "2026-06-03T18:04:00.000", "2026-06-03T18:04:00.020", "2026-06-03T18:04:00.040",
        ]
        # The original Date/Time source columns never appear.
        assert "Date" not in rows[0]

    def test_original_source_never_mutated(self):
        prep = PreparationSessionRegistry()
        content = b"2026-08-31 13:00:00,1.0\n"
        sid = _ready_absolute_source(prep, content)

        export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        session = prep.get(WS, sid)
        assert session.raw_bytes == content

    def test_reported_minute_resolution_exports_zero_seconds(self):
        # Enhancement (minute/AM-PM-hour absolute time support): the
        # exact owner-reported example -- task section V's own worked
        # example.
        prep = PreparationSessionRegistry()
        content = b"3/6/2026 17:25,1.0\n3/6/2026 17:26,2.0\n3/6/2026 17:27,3.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0, date_order="dmy")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert [r[0] for r in rows[1:]] == [
            "2026-06-03T17:25:00.000", "2026-06-03T17:26:00.000", "2026-06-03T17:27:00.000",
        ]

    def test_explicit_am_pm_hour_only_exports_canonical_24h(self):
        # Task section W's own worked example.
        prep = PreparationSessionRegistry()
        content = b"2026-06-03 1pm,1.0\n2026-06-03 2pm,2.0\n"
        sid = _ready_absolute_source(prep, content)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert [r[0] for r in rows[1:]] == ["2026-06-03T13:00:00.000", "2026-06-03T14:00:00.000"]


class TestTimezoneExport:
    def test_real_offset_preserved(self):
        prep = PreparationSessionRegistry()
        content = b"2026-06-03T18:04:00.000+08:00,1.0\n2026-06-03T18:04:00.020+08:00,2.0\n"
        sid = _ready_absolute_source(prep, content)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))
        manifest = _read_manifest(_unzip(result.content))

        assert [r[0] for r in rows[1:]] == ["2026-06-03T18:04:00.000+08:00", "2026-06-03T18:04:00.020+08:00"]
        assert manifest["exported_time"]["timezone_present"] is True

    def test_absent_timezone_never_invented(self):
        prep = PreparationSessionRegistry()
        content = b"2026-06-03 18:04:00,1.0\n"
        sid = _ready_absolute_source(prep, content)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))
        manifest = _read_manifest(_unzip(result.content))

        exported_value = rows[1][0]
        assert exported_value == "2026-06-03T18:04:00.000"
        assert "Z" not in exported_value
        assert "+" not in exported_value
        assert manifest["exported_time"]["timezone_present"] is False


class TestPrecisionExport:
    def test_millisecond_precision_preserved(self):
        prep = PreparationSessionRegistry()
        content = b"2026-06-03 18:04:00.020,1.0\n"
        sid = _ready_absolute_source(prep, content)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert rows[1][0] == "2026-06-03T18:04:00.020"

    def test_sub_millisecond_precision_preserved(self):
        prep = PreparationSessionRegistry()
        content = b"2026-06-03 18:04:00.123456,1.0\n"
        sid = _ready_absolute_source(prep, content)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert rows[1][0] == "2026-06-03T18:04:00.123456"


class TestElapsedTimeExport:
    def test_relative_to_first_active_row(self):
        # Task section AE's own worked example.
        prep = PreparationSessionRegistry()
        content = b"5.000,1.0\n5.020,2.0\n5.040,3.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_elapsed(prep, sid, unit="seconds")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))
        manifest = _read_manifest(_unzip(result.content))

        assert rows[0] == ["Time (s)", "B"]
        assert [r[0] for r in rows[1:]] == ["0.000", "0.020", "0.040"]
        assert manifest["exported_time"]["source_offset_seconds"] == 5.0
        assert manifest["exported_time"]["family"] == "partial" or manifest["exported_time"]["family"] == "elapsed"

    def test_milliseconds_unit_converted_to_seconds(self):
        prep = PreparationSessionRegistry()
        content = b"5000,1.0\n5020,2.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_elapsed(prep, sid, unit="milliseconds")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert [r[0] for r in rows[1:]] == ["0.000", "0.020"]

    @pytest.mark.parametrize(
        "unit, values, expected",
        [
            ("minutes", ["0", "1", "2"], ["0.000", "60.000", "120.000"]),
            ("hours", ["0", "1", "2"], ["0.000", "3600.000", "7200.000"]),
            ("days", ["0", "1"], ["0.000", "86400.000"]),
            ("weeks", ["0", "0.5"], ["0.000", "302400.000"]),
        ],
    )
    def test_fixed_duration_units_normalized_to_seconds(self, unit, values, expected):
        prep = PreparationSessionRegistry()
        content = "".join(f"{v},{i}.0\n" for i, v in enumerate(values)).encode()
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_elapsed(prep, sid, unit=unit)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert [r[0] for r in rows[1:]] == expected


class TestSampleIndexTimeExport:
    def test_known_interval_produces_relative_seconds(self):
        # Task section AF's own worked example.
        prep = PreparationSessionRegistry()
        content = b"100,1.0\n101,2.0\n102,3.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_sample_index(prep, sid, interval_seconds=0.02)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))
        manifest = _read_manifest(_unzip(result.content))

        assert rows[0] == ["Time (s)", "B"]
        assert [r[0] for r in rows[1:]] == ["0.000", "0.020", "0.040"]
        assert manifest["exported_time"]["interval_seconds"] == 0.02

    def test_missing_interval_refuses_export_with_clear_message(self):
        prep = PreparationSessionRegistry()
        content = b"100,1.0\n101,2.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_sample_index(prep, sid, interval_seconds=None)

        with pytest.raises(ExportRequiresIntervalError) as excinfo:
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        assert "interval" in str(excinfo.value.message).lower()


class TestReconstructedTimeExport:
    def test_accepted_reconstruction_exports_resolved_time(self):
        # Task section AG's own worked example: a coarse, repeated-
        # every-other-row timestamp (2 samples per second) reconstructed
        # to a real sub-second cadence.
        prep = PreparationSessionRegistry()
        lines = [f"2026-08-31 13:00:{i // 2:02d},{i}.0" for i in range(6)]
        content = ("\n".join(lines) + "\n").encode()
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        preview = interpret_time_axis(
            workspace_id=WS, source_id=sid, column_indices=(0,),
            interpreter_id="repeated_timestamp_precision_loss", registry=prep,
        )
        assert preview.resolved_interval_seconds is not None
        set_time_axis_configuration(
            workspace_id=WS, source_id=sid, column_indices=(0,),
            interpreter_id="repeated_timestamp_precision_loss", confirmed=True, registry=prep,
        )

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))
        manifest = _read_manifest(_unzip(result.content))

        # The original coarse (all-identical) timestamp is NOT what
        # ends up in the export -- the reconstructed cadence is.
        exported_times = [r[0] for r in rows[1:]]
        assert len(set(exported_times)) == len(exported_times)  # no longer all-identical
        assert manifest["exported_time"]["reconstructed"] is True


class TestPartialTimeExport:
    def test_time_only_column_exports_relative_seconds_no_fabricated_date(self):
        prep = PreparationSessionRegistry()
        content = b"18:04:00.000,1.0\n18:04:00.020,2.0\n18:04:00.040,3.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        # Explicit interpreter authority: this is genuinely time-of-day
        # data, so the correct (non-mismatched) explicit interpreter is
        # `time_of_day` -- selecting `absolute_datetime` for a bare
        # clock-time column is now the exact scenario the interpreter/
        # family compatibility guard flags, not a valid way to reach
        # FAMILY_PARTIAL.
        set_time_axis_configuration(
            workspace_id=WS, source_id=sid, column_indices=(0,),
            interpreter_id="time_of_day", confirmed=True, registry=prep,
        )

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))
        manifest = _read_manifest(_unzip(result.content))

        assert rows[0] == ["Time (s)", "B"]
        assert [r[0] for r in rows[1:]] == ["0.000", "0.020", "0.040"]
        assert manifest["exported_time"]["family"] == "partial"
        assert not any("-" in v and "T" in v for v in [r[0] for r in rows[1:]])  # no fabricated ISO date


class TestWaveformDataIntegrity:
    def test_working_edits_reflected(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,ERR\n")
        edit_cell(workspace_id=WS, source_id=sid, row_number=1, column_index=1, value="2.5", registry=prep)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert rows[1][1] == "2.5"

    def test_clearing_a_required_waveform_cell_now_blocks_export(self):
        # Under the new gate (task section AI: "if current readiness
        # blocks missing waveform values, keep that existing policy"),
        # a cleared cell in an otherwise-active Waveform column is
        # exactly the SAME "missing value" readiness already blocks --
        # never silently exported as an empty field.
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1\n")
        edit_cell(workspace_id=WS, source_id=sid, row_number=1, column_index=1, value=None, registry=prep)

        with pytest.raises(ExportNotReadyError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

    def test_cleared_cell_in_an_excluded_row_never_blocks_export(self):
        # The one legitimate "cleared and still exportable" case: the
        # row containing the cleared cell is itself excluded (readiness
        # never scans excluded rows), so the whole row -- cleared cell
        # included -- simply never reaches the export table at all.
        prep = PreparationSessionRegistry()
        content = b"2026-08-31 13:00:00,1\n2026-08-31 13:00:01,2\n"
        sid = _ready_absolute_source(prep, content)
        edit_cell(workspace_id=WS, source_id=sid, row_number=1, column_index=1, value=None, registry=prep)
        set_row_excluded(workspace_id=WS, source_id=sid, row_number=1, excluded=True, registry=prep)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert len(rows) == 2  # header + the one remaining (non-cleared) row
        assert rows[1][1] == "2"

    def test_excluded_rows_omitted_from_both_time_and_waveform(self):
        prep = PreparationSessionRegistry()
        content = b"2026-08-31 13:00:00,1\n2026-08-31 13:00:01,2\n2026-08-31 13:00:02,3\n"
        sid = _ready_absolute_source(prep, content)
        set_row_excluded(workspace_id=WS, source_id=sid, row_number=2, excluded=True, registry=prep)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert len(rows) == 3  # header + 2 data rows
        assert [r[1] for r in rows[1:]] == ["1", "3"]

    def test_not_assigned_columns_omitted(self):
        prep = PreparationSessionRegistry()
        content = b"Time,Voltage,Status\n2026-08-31 13:00:00,1.0,ok\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=2, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        # column_index=2 ("Status") is left at its default (not_assigned)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))
        manifest = _read_manifest(_unzip(result.content))

        assert rows[0] == ["Time", "Voltage"]
        assert manifest["omitted_columns"] == [{"column_index": 2, "label": "Status", "role": "not_assigned"}]

    def test_one_configured_time_value_per_exported_row(self):
        prep = PreparationSessionRegistry()
        lines = [f"2026-08-31 13:00:{i:02d},{i}.0" for i in range(10)]
        content = ("\n".join(lines) + "\n").encode()
        sid = _ready_absolute_source(prep, content)
        set_row_excluded(workspace_id=WS, source_id=sid, row_number=3, excluded=True, registry=prep)
        set_row_excluded(workspace_id=WS, source_id=sid, row_number=7, excluded=True, registry=prep)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert len(rows) - 1 == 8  # 10 rows - 2 excluded

    def test_waveform_column_source_order_preserved(self):
        prep = PreparationSessionRegistry()
        content = b"2026-08-31 13:00:00,10.0,20.0,30.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1, 2, 3)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert rows[0] == ["Time", "B", "C", "D"]
        assert rows[1] == ["2026-08-31T13:00:00.000", "10.0", "20.0", "30.0"]

    def test_configured_time_column_always_first_regardless_of_source_position(self):
        # Task section R's own worked example: Voltage | Date | Current | Time
        prep = PreparationSessionRegistry()
        content = b"Voltage,Date,Current,Time\n1.0,2026-08-31,2.0,13:00:00\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=2, registry=prep)
        _mark_waveform(prep, sid, 0, 2)
        _mark_time_axis(prep, sid, 1, 3)
        _confirm_split_date_time(prep, sid, date_column=1, time_column=3, date_order="ymd")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert rows[0] == ["Time", "Voltage", "Current"]

    def test_duplicate_waveform_labels_survive_uniquely(self):
        prep = PreparationSessionRegistry()
        content = b"Time,Voltage,Voltage\n2026-08-31 13:00:00,1,2\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=2, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1, 2)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))

        assert rows[0] == ["Time", "Voltage", "Voltage__C"]


class TestManifestExportedTime:
    def test_manifest_records_source_columns_and_provenance(self):
        prep = PreparationSessionRegistry()
        content = b"Date,Time,Voltage\n3/6/26,18:04:00.000,1.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=2, registry=prep)
        _mark_time_axis(prep, sid, 0, 1)
        _mark_waveform(prep, sid, 2)
        _confirm_split_date_time(prep, sid, date_column=0, time_column=1, date_order="dmy")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        manifest = _read_manifest(_unzip(result.content))
        exported_time = manifest["exported_time"]

        assert exported_time["column_name"] == "Time"
        assert exported_time["source_columns"] == [
            {"column_index": 0, "label": "Date"}, {"column_index": 1, "label": "Time"},
        ]
        assert exported_time["family"] == "absolute"
        assert exported_time["interpreter_id"] == "split_date_time"
        assert exported_time["date_order"] == "dmy"
        assert exported_time["export_representation"] == "iso8601"
        assert exported_time["reconstructed"] is False

    def test_existing_slice_12_provenance_fields_retained(self):
        prep = PreparationSessionRegistry()
        content = b"Time,Voltage\n2026-08-31 13:00:00,1\n2026-08-31 13:00:01,2\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        manifest = _read_manifest(_unzip(result.content))

        for key in (
            "manifest_version", "exported_at", "exported_file", "source_format", "original_filename",
            "worksheet_name", "worksheet_index", "preparation_revision", "header_row", "data_region",
            "exported_row_count", "excluded_row_count", "excluded_rows", "excluded_rows_truncated",
            "omitted_columns", "column_roles", "edited_cell_count", "cleared_cell_count",
            "time_family", "time_provenance", "interpreter_id", "time_unit", "time_interval_seconds",
            "reconstructed_timing", "exported_time", "readiness",
        ):
            assert key in manifest, key
        assert set(manifest["readiness"].keys()) == {"is_ready", "blocking_count", "warning_count", "info_count"}
        assert manifest["readiness"]["is_ready"] is True


class TestExcelExport:
    def test_configured_time_and_waveform_only(self):
        prep = PreparationSessionRegistry()
        content = _build_xlsx({
            "A": [["Time", "VR"], ["2026-08-31 13:00:00", 1.0], ["2026-08-31 13:00:01", 2.0]],
        })
        sid = _add_excel(prep, content)
        select_preparation_worksheet(workspace_id=WS, source_id=sid, worksheet_index=0, registry=prep)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        assert result.filename.endswith(".zip")
        rows = _read_xlsx_rows(_unzip(result.content))

        assert rows[0] == ("Time", "VR")
        assert rows[1] == ("2026-08-31T13:00:00.000", 1.0)

    def test_output_is_a_single_clean_worksheet(self):
        prep = PreparationSessionRegistry()
        content = _build_xlsx({"Only": [["2026-08-31 13:00:00", "1"]]})
        sid = _add_excel(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        name = next(n for n in _unzip(result.content).namelist() if n.endswith(".xlsx"))
        wb = load_workbook(io.BytesIO(_unzip(result.content).read(name)))

        assert len(wb.sheetnames) == 1

    def test_original_workbook_bytes_untouched(self):
        prep = PreparationSessionRegistry()
        content = _build_xlsx({"Sheet1": [["2026-08-31 13:00:00", "1"]]})
        sid = _add_excel(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        session = prep.get(WS, sid)
        assert session.raw_bytes == content

    def test_valid_worksheet_name_preserved(self):
        prep = PreparationSessionRegistry()
        content = _build_xlsx({"Event Data": [["2026-08-31 13:00:00", "1"]]})
        sid = _add_excel(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        name = next(n for n in _unzip(result.content).namelist() if n.endswith(".xlsx"))
        wb = load_workbook(io.BytesIO(_unzip(result.content).read(name)))

        assert wb.sheetnames == ["Event Data"]


class TestExportIsReadOnly:
    def test_revision_unchanged_by_export(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1\n")
        session = prep.get(WS, sid)
        before = session.working_overlay.revision

        export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        assert session.working_overlay.revision == before

    def test_preparation_session_still_present_after_export(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1\n")

        export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        assert prep.get(WS, sid) is not None

    def test_repeated_export_is_idempotent_and_consistent(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1\n")

        first = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        second = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        rows_first = _read_csv_rows(_unzip(first.content))
        rows_second = _read_csv_rows(_unzip(second.content))
        assert rows_first == rows_second
        assert prep.get(WS, sid) is not None


class TestApiLevelErrors:
    def test_missing_source_raises_source_not_found(self):
        prep = PreparationSessionRegistry()
        with pytest.raises(SourceNotFoundError):
            export_preparation_source(workspace_id=WS, source_id="does-not-exist", registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

    def test_excel_worksheet_not_selected_raises(self):
        prep = PreparationSessionRegistry()
        content = _build_xlsx({"A": [["a"]], "B": [["b"]]})
        sid = _add_excel(prep, content)
        with pytest.raises(WorksheetNotSelectedError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)


class TestFilenameSanitization:
    def test_csv_filename_pattern(self):
        prep = PreparationSessionRegistry()
        sid = _add_csv(prep, b"2026-08-31 13:00:00,1\n", filename="event.csv")
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        assert result.filename == "event_cleaned.zip"
        names = _unzip(result.content).namelist()
        assert "event_cleaned.csv" in names
        assert "event_cleaned.manifest.json" in names

    def test_excel_filename_pattern(self):
        prep = PreparationSessionRegistry()
        content = _build_xlsx({"Sheet1": [["2026-08-31 13:00:00", "1"]]})
        sid = _add_excel(prep, content, filename="event.xlsx")
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        assert result.filename == "event_cleaned.zip"
        names = _unzip(result.content).namelist()
        assert "event_cleaned.xlsx" in names

    def test_unsafe_characters_in_original_filename_stripped(self):
        prep = PreparationSessionRegistry()
        sid = _add_csv(prep, b"2026-08-31 13:00:00,1\n", filename="../weird:name*.csv")
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        assert "/" not in result.filename
        assert ":" not in result.filename
        assert "*" not in result.filename


class TestPerformanceLargeExport:
    def test_large_csv_exports_without_pathological_cost(self):
        prep = PreparationSessionRegistry()
        rows = 20_000
        lines = [f"2026-08-31 13:00:{i % 60:02d}.{i:06d},{float(i)}" for i in range(rows)]
        sid = _add_csv(prep, ("\n".join(lines) + "\n").encode())
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows_out = _read_csv_rows(_unzip(result.content))

        assert len(rows_out) == rows + 1  # header + all rows

    def test_large_excel_exports_via_write_only_streaming(self):
        import datetime as _dt

        prep = PreparationSessionRegistry()
        rows = 10_000
        base = _dt.datetime(2026, 8, 31, 13, 0, 0)
        content = _build_xlsx({
            "Sheet1": [[(base + _dt.timedelta(seconds=i)).strftime("%Y-%m-%d %H:%M:%S"), float(i)] for i in range(rows)],
        })
        sid = _add_excel(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows_out = _read_xlsx_rows(_unzip(result.content))

        assert len(rows_out) == rows + 1  # header + all rows


class TestReUploadRoundTrip:
    """Task section AK: a cleaned export's own standardized Time column
    must be recognizable on re-upload without repeating the original
    ambiguity/configuration work."""

    def test_absolute_export_recognized_unambiguously_on_reupload(self):
        prep = PreparationSessionRegistry()
        # Originally ambiguous (DMY vs MDY) -- resolved once, exported.
        content = b"Date,Time,Voltage\n3/6/26,18:04:00.000,1.0\n3/6/26,18:04:00.020,2.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0, 1)
        _mark_waveform(prep, sid, 2)
        _confirm_split_date_time(prep, sid, date_column=0, time_column=1, date_order="dmy")
        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        csv_bytes = _unzip(result.content).read(next(n for n in _unzip(result.content).namelist() if n.endswith(".csv")))

        # Re-upload the cleaned CSV as a brand-new preparation source.
        sid2 = _add_csv(prep, csv_bytes, filename="reuploaded.csv")
        set_header_row(workspace_id=WS, source_id=sid2, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid2, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid2, 0)
        preview = interpret_time_axis(
            workspace_id=WS, source_id=sid2, column_indices=(0,), interpreter_id="absolute_datetime", registry=prep,
        )

        # No ambiguity remains -- the standardized ISO column is
        # self-describing, resolved without needing a date_order choice.
        assert preview.family == "absolute"
        assert not any(d.code == "ambiguous_date_order" for d in preview.diagnostics)

    def test_elapsed_export_recognized_on_reupload(self):
        prep = PreparationSessionRegistry()
        content = b"5.000,1.0\n5.020,2.0\n5.040,3.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_elapsed(prep, sid, unit="seconds")
        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        csv_bytes = _unzip(result.content).read(next(n for n in _unzip(result.content).namelist() if n.endswith(".csv")))

        sid2 = _add_csv(prep, csv_bytes, filename="reuploaded.csv")
        set_header_row(workspace_id=WS, source_id=sid2, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid2, start_row=2, end_row=4, registry=prep)
        _mark_time_axis(prep, sid2, 0)
        preview = interpret_time_axis(
            workspace_id=WS, source_id=sid2, column_indices=(0,), interpreter_id="elapsed_numeric",
            unit="seconds", registry=prep,
        )

        assert preview.family == "elapsed"
        assert not any(d.ambiguity == "ambiguous" for d in preview.diagnostics)

    def test_sample_index_export_recognized_on_reupload(self):
        prep = PreparationSessionRegistry()
        content = b"100,1.0\n101,2.0\n102,3.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_sample_index(prep, sid, interval_seconds=0.02)
        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        csv_bytes = _unzip(result.content).read(next(n for n in _unzip(result.content).namelist() if n.endswith(".csv")))

        sid2 = _add_csv(prep, csv_bytes, filename="reuploaded.csv")
        set_header_row(workspace_id=WS, source_id=sid2, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid2, start_row=2, end_row=4, registry=prep)
        _mark_time_axis(prep, sid2, 0)
        preview = interpret_time_axis(
            workspace_id=WS, source_id=sid2, column_indices=(0,), interpreter_id="elapsed_numeric",
            unit="seconds", registry=prep,
        )

        # The re-uploaded "Time (s)" column is itself a plain elapsed-
        # seconds column -- recognized immediately by elapsed_numeric,
        # no interval/rate re-entry required (unlike the ORIGINAL
        # sample-index source, which had none at all).
        assert preview.family == "elapsed"

    def test_minute_resolution_export_recognized_unambiguously_on_reupload(self):
        # Enhancement (minute/AM-PM-hour absolute time support), task
        # section AG: the exact owner-reported minute-resolution example,
        # once resolved and exported, must be self-describing on re-upload.
        prep = PreparationSessionRegistry()
        content = b"3/6/2026 17:25,1.0\n3/6/2026 17:26,2.0\n3/6/2026 17:27,3.0\n"
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_absolute(prep, sid, column_index=0, date_order="dmy")
        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        csv_bytes = _read_csv_rows(_unzip(result.content))
        assert [r[0] for r in csv_bytes[1:]] == [
            "2026-06-03T17:25:00.000", "2026-06-03T17:26:00.000", "2026-06-03T17:27:00.000",
        ]
        raw_csv_bytes = _unzip(result.content).read(next(n for n in _unzip(result.content).namelist() if n.endswith(".csv")))

        sid2 = _add_csv(prep, raw_csv_bytes, filename="reuploaded.csv")
        set_header_row(workspace_id=WS, source_id=sid2, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid2, start_row=2, end_row=4, registry=prep)
        _mark_time_axis(prep, sid2, 0)
        preview = interpret_time_axis(
            workspace_id=WS, source_id=sid2, column_indices=(0,), interpreter_id="absolute_datetime", registry=prep,
        )

        assert preview.family == "absolute"
        assert not any(d.code == "ambiguous_date_order" for d in preview.diagnostics)

    def test_am_pm_hour_only_export_recognized_on_reupload(self):
        prep = PreparationSessionRegistry()
        content = b"2026-06-03 1pm,1.0\n2026-06-03 2pm,2.0\n"
        sid = _ready_absolute_source(prep, content)
        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        rows = _read_csv_rows(_unzip(result.content))
        assert [r[0] for r in rows[1:]] == ["2026-06-03T13:00:00.000", "2026-06-03T14:00:00.000"]
        raw_csv_bytes = _unzip(result.content).read(next(n for n in _unzip(result.content).namelist() if n.endswith(".csv")))

        sid2 = _add_csv(prep, raw_csv_bytes, filename="reuploaded.csv")
        set_header_row(workspace_id=WS, source_id=sid2, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid2, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid2, 0)
        preview = interpret_time_axis(
            workspace_id=WS, source_id=sid2, column_indices=(0,), interpreter_id="absolute_datetime", registry=prep,
        )

        assert preview.family == "absolute"
        assert not any(d.code == "ambiguous_date_order" for d in preview.diagnostics)

    @pytest.mark.parametrize("unit, values", [
        ("minutes", ["0", "1", "2"]),
        ("hours", ["0", "1", "2"]),
        ("days", ["0", "1"]),
        ("weeks", ["0", "0.5"]),
    ])
    def test_fixed_duration_elapsed_units_export_recognized_on_reupload(self, unit, values):
        prep = PreparationSessionRegistry()
        content = "".join(f"{v},{i}.0\n" for i, v in enumerate(values)).encode()
        sid = _add_csv(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_elapsed(prep, sid, unit=unit)
        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)
        raw_csv_bytes = _unzip(result.content).read(next(n for n in _unzip(result.content).namelist() if n.endswith(".csv")))

        sid2 = _add_csv(prep, raw_csv_bytes, filename="reuploaded.csv")
        set_header_row(workspace_id=WS, source_id=sid2, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid2, start_row=2, end_row=1 + len(values), registry=prep)
        _mark_time_axis(prep, sid2, 0)
        preview = interpret_time_axis(
            workspace_id=WS, source_id=sid2, column_indices=(0,), interpreter_id="elapsed_numeric",
            unit="seconds", registry=prep,
        )

        # The re-exported "Time (s)" column is already normalized to
        # plain elapsed seconds -- the original unit choice need not be
        # remembered or re-entered on re-upload.
        assert preview.family == "elapsed"
        assert not any(d.ambiguity == "ambiguous" for d in preview.diagnostics)


# ---- 2026-09-04 UAT enhancement: manifest/provenance is optional ----


def _read_xlsx_rows_direct(content: bytes) -> list[tuple]:
    """Like `_read_xlsx_rows()` above, but for a DATA-ONLY export result
    -- `content` is the raw XLSX workbook bytes themselves, never a ZIP
    member."""
    return list(load_workbook(io.BytesIO(content)).active.iter_rows(values_only=True))


class TestDataOnlyExport:
    """`EXPORT_MODE_DATA_ONLY` is the new DEFAULT -- the plain "Export
    Cleaned Data" action returns the cleaned CSV/XLSX bytes directly,
    never a ZIP, never a manifest. Provenance capability itself is not
    removed (see `TestExportGating`/every other class above, all still
    exercising `EXPORT_MODE_WITH_PROVENANCE` unchanged) -- only the
    DEFAULT shape changes."""

    def test_default_mode_is_data_only(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1.0\n")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep)

        assert result.filename == "e_cleaned.csv"
        assert result.media_type == "text/csv"

    def test_csv_default_returns_csv_not_zip(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1.0\n2026-08-31 13:00:01,2.0\n")

        result = export_preparation_source(
            workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY,
        )

        assert result.filename == "e_cleaned.csv"
        assert result.media_type == "text/csv"
        with pytest.raises(zipfile.BadZipFile):
            _unzip(result.content)
        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        # No header row is configured by `_ready_absolute_source()` --
        # the Waveform column's own label falls back to its plain
        # spreadsheet letter, same as every other no-header case in this
        # file (see e.g. TestWaveformIntegrity's own `["Time (s)", "B"]`).
        assert rows[0] == ["Time", "B"]
        assert rows[1][0] == "2026-08-31T13:00:00.000"
        assert rows[2][0] == "2026-08-31T13:00:01.000"

    def test_excel_default_returns_xlsx_not_zip(self):
        prep = PreparationSessionRegistry()
        content = _build_xlsx({"Sheet1": [["Time", "VR"], ["2026-08-31 13:00:00", 1.0], ["2026-08-31 13:00:01", 2.0]]})
        sid = _add_excel(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _confirm_absolute(prep, sid, column_index=0)

        result = export_preparation_source(
            workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY,
        )

        assert result.filename == "e_cleaned.xlsx"
        assert result.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # An XLSX is technically a ZIP container internally, so the real
        # proof of "export/HTTP semantics, not a provenance bundle" is
        # the filename/media_type above (never `.zip`/`application/zip`)
        # plus the absence of a manifest member below -- never an
        # assertion that the bytes aren't ZIP-formatted at all.
        assert not any(n.endswith(".manifest.json") for n in zipfile.ZipFile(io.BytesIO(result.content)).namelist())
        rows = _read_xlsx_rows_direct(result.content)
        assert rows[0] == ("Time", "VR")
        assert rows[1] == ("2026-08-31T13:00:00.000", 1.0)

    def test_data_only_export_gated_identically_to_with_provenance(self):
        # Same gating fixture as TestExportGating.test_unconfigured_time_axis_blocks_export.
        prep = PreparationSessionRegistry()
        sid = _add_csv(prep, b"1,2\n")
        _mark_waveform(prep, sid, 1)

        with pytest.raises(ExportNotReadyError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

    def test_data_only_export_requires_interval_gate_identical(self):
        prep = PreparationSessionRegistry()
        sid = _add_csv(prep, b"100,1.0\n101,2.0\n")
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_sample_index(prep, sid, interval_seconds=None)

        with pytest.raises(ExportRequiresIntervalError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

    def test_data_only_export_preserves_preparation_state(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1.0\n")
        before_revision = prep.get(WS, sid).working_overlay.revision

        export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        after = prep.get(WS, sid)
        assert after.working_overlay.revision == before_revision
        assert after.raw_bytes == b"2026-08-31 13:00:00,1.0\n"

    def test_unknown_mode_rejected(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1.0\n")

        with pytest.raises(ValueError):
            export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode="not-a-real-mode")


class TestModeEquivalence:
    """Task section J: both export modes must contain the exact same
    cleaned data -- `mode` only changes the return SHAPE (direct bytes
    vs. ZIP+manifest), never the cleaned-data construction itself (both
    share `_ensure_exportable()`/the same row-assembly code path)."""

    def test_csv_data_only_matches_with_provenance_bundle(self):
        prep = PreparationSessionRegistry()
        content = b"Date,Time,Voltage\n3/6/26,18:04:00.000,132.1\n3/6/26,18:04:00.020,132.2\n3/6/26,18:04:00.040,132.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=4, registry=prep)
        _mark_time_axis(prep, sid, 0, 1)
        _mark_waveform(prep, sid, 2)
        _confirm_split_date_time(prep, sid, date_column=0, time_column=1, date_order="dmy")

        data_only = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)
        with_provenance = export_preparation_source(
            workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE,
        )

        bundled_csv_bytes = _unzip(with_provenance.content).read(
            next(n for n in _unzip(with_provenance.content).namelist() if n.endswith(".csv"))
        )
        assert data_only.content == bundled_csv_bytes

    def test_excel_data_only_matches_with_provenance_bundle(self):
        prep = PreparationSessionRegistry()
        content = _build_xlsx({"Sheet1": [["Time", "VR"], ["2026-08-31 13:00:00", 1.0], ["2026-08-31 13:00:01", 2.0]]})
        sid = _add_excel(prep, content)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _confirm_absolute(prep, sid, column_index=0)

        data_only = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)
        with_provenance = export_preparation_source(
            workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE,
        )

        bundled_xlsx_bytes = _unzip(with_provenance.content).read(
            next(n for n in _unzip(with_provenance.content).namelist() if n.endswith(".xlsx"))
        )
        assert _read_xlsx_rows_direct(data_only.content) == _read_xlsx_rows_direct(bundled_xlsx_bytes)

    def test_filenames_differ_by_mode_same_base_name(self):
        prep = PreparationSessionRegistry()
        sid = _ready_absolute_source(prep, b"2026-08-31 13:00:00,1.0\n")

        data_only = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)
        with_provenance = export_preparation_source(
            workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE,
        )

        assert data_only.filename == "e_cleaned.csv"
        assert with_provenance.filename == "e_cleaned.zip"


class TestEngineeringQuantityExportLabels:
    """DEC-077 task sections N-U, AK-AM: cleaned CSV/XLSX waveform column
    labels encode the Engineering Quantity as a strict suffix; the
    optional manifest may continue recording it too, but restoration
    never depends on the manifest."""

    def _labeled_source(self, prep, *, quantity: str | None) -> str:
        content = b"Time,CBDK_V1 Magnitude\n2026-08-31 13:00:00,1.0\n2026-08-31 13:00:01,2.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        if quantity is not None:
            set_column_engineering_quantity(
                workspace_id=WS, source_id=sid, column_index=1, engineering_quantity=quantity, registry=prep,
            )
        _confirm_absolute(prep, sid, column_index=0)
        return sid

    def test_known_quantity_appends_suffix_to_header(self):
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity="Voltage")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        assert rows[0] == ["Time", "CBDK_V1 Magnitude (Voltage)"]

    @pytest.mark.parametrize(
        "quantity,expected_suffix",
        [
            ("Voltage", "(Voltage)"),
            ("Voltage Angle", "(Voltage Angle)"),
            ("Current", "(Current)"),
            ("Current Angle", "(Current Angle)"),
            ("Active Power", "(Active Power)"),
            ("Reactive Power", "(Reactive Power)"),
            ("Frequency", "(Frequency)"),
            ("ROCOF", "(ROCOF)"),
        ],
    )
    def test_every_known_quantity_produces_its_own_exact_suffix(self, quantity, expected_suffix):
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity=quantity)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        assert rows[0][1] == f"CBDK_V1 Magnitude {expected_suffix}"

    def test_undefined_quantity_never_gets_a_suffix(self):
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity=None)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        assert rows[0] == ["Time", "CBDK_V1 Magnitude"]
        assert "(Undefined)" not in rows[0][1]

    def test_configured_time_column_header_never_gets_a_quantity_suffix(self):
        # Task section T: the Configured Time column's own "(s)"-style
        # header is never confused with an Engineering Quantity suffix --
        # it is produced entirely separately (_build_configured_time_column)
        # and never passed through encode_engineering_quantity_suffix().
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity="Voltage")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        assert rows[0][0] == "Time"

    def test_manifest_may_still_record_engineering_quantities(self):
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity="Voltage")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        manifest = _read_manifest(_unzip(result.content))
        assert manifest["column_engineering_quantities"][1] == "Voltage"

    def test_data_only_export_needs_no_manifest_to_carry_the_quantity(self):
        # The header suffix alone is the carrier -- EXPORT_MODE_DATA_ONLY
        # builds no manifest at all, and the suffix is still present.
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity="ROCOF")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        with pytest.raises(zipfile.BadZipFile):
            _unzip(result.content)
        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        assert rows[0][1] == "CBDK_V1 Magnitude (ROCOF)"


class TestEngineeringQuantityReUploadRoundTrip:
    """DEC-077 task sections S, AL: re-uploading a cleaned export restores
    Engineering Quantity deterministically from the header suffix, and
    exporting again is stable (never a duplicated suffix)."""

    def test_full_round_trip_restores_and_re_exports_identically(self):
        prep = PreparationSessionRegistry()
        first_sid = self._first_export_sid(prep)
        first_result = export_preparation_source(
            workspace_id=WS, source_id=first_sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY,
        )
        assert first_result.content.decode("utf-8").splitlines()[0] == "Time,CBDK_V1 Magnitude (Voltage)"

        # Re-upload the cleaned export as a brand-new preparation source.
        second_sid = _add_csv(prep, first_result.content, filename="re-uploaded.csv")
        set_header_row(workspace_id=WS, source_id=second_sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=second_sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, second_sid, 0)
        _mark_waveform(prep, second_sid, 1)  # restoration fires here
        _confirm_absolute(prep, second_sid, column_index=0)

        session = prep.get(WS, second_sid)
        assert session.working_overlay.column_engineering_quantities[column_key(None, 1)] == "Voltage"

        second_result = export_preparation_source(
            workspace_id=WS, source_id=second_sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY,
        )
        second_header = second_result.content.decode("utf-8").splitlines()[0]
        # Stable, never duplicated -- "... (Voltage) (Voltage)" would be a bug.
        assert second_header == "Time,CBDK_V1 Magnitude (Voltage)"

    def _first_export_sid(self, prep) -> str:
        content = b"Time,CBDK_V1 Magnitude\n2026-08-31 13:00:00,1.0\n2026-08-31 13:00:01,2.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        set_column_engineering_quantity(
            workspace_id=WS, source_id=sid, column_index=1, engineering_quantity="Voltage", registry=prep,
        )
        _confirm_absolute(prep, sid, column_index=0)
        return sid

    def test_time_header_suffix_never_confused_with_engineering_quantity_on_reupload(self):
        # "Time (s)" (a non-absolute Configured Time header) must never
        # be misparsed as an Engineering Quantity suffix on re-upload.
        prep = PreparationSessionRegistry()
        content = b"Elapsed,Value\n0.0,1.0\n0.02,2.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        _confirm_elapsed(prep, sid, column_index=0)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)
        assert result.content.decode("utf-8").splitlines()[0] == "Time (s),Value"

        second_sid = _add_csv(prep, result.content, filename="re-uploaded.csv")
        set_header_row(workspace_id=WS, source_id=second_sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=second_sid, start_row=2, end_row=3, registry=prep)
        _mark_waveform(prep, second_sid, 0)  # "Time (s)" column, misassigned Waveform on purpose

        session = prep.get(WS, second_sid)
        assert column_key(None, 0) not in session.working_overlay.column_engineering_quantities


class TestMeasuredUnitExportLabels:
    """Measured Unit enhancement (DEC-080), task section P/Q/AP: cleaned
    CSV/XLSX waveform column labels encode Measured Unit as an
    additional strict suffix after Engineering Quantity."""

    def _labeled_source(self, prep, *, quantity: str | None, unit: str | None) -> str:
        content = b"Time,CBDK_V1 Magnitude\n2026-08-31 13:00:00,1.0\n2026-08-31 13:00:01,2.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        if quantity is not None:
            set_column_engineering_quantity(
                workspace_id=WS, source_id=sid, column_index=1, engineering_quantity=quantity, registry=prep,
            )
        if unit is not None:
            set_column_measured_unit(
                workspace_id=WS, source_id=sid, column_index=1, measured_unit=unit, registry=prep,
            )
        _confirm_absolute(prep, sid, column_index=0)
        return sid

    @pytest.mark.parametrize(
        "quantity,unit,expected_suffix",
        [
            ("Voltage", "kV", "(Voltage) [kV]"),
            ("Voltage Angle", "deg", "(Voltage Angle) [deg]"),
            ("Current", "A", "(Current) [A]"),
            ("Current Angle", "deg", "(Current Angle) [deg]"),
            ("Active Power", "MW", "(Active Power) [MW]"),
            ("Reactive Power", "Mvar", "(Reactive Power) [Mvar]"),
            ("Frequency", "Hz", "(Frequency) [Hz]"),
            ("ROCOF", "Hz/s", "(ROCOF) [Hz/s]"),
        ],
    )
    def test_every_known_quantity_and_unit_produces_its_own_exact_suffix(self, quantity, unit, expected_suffix):
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity=quantity, unit=unit)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        assert rows[0][1] == f"CBDK_V1 Magnitude {expected_suffix}"

    def test_blank_unit_omits_bracket(self):
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity="Voltage", unit=None)

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        assert rows[0][1] == "CBDK_V1 Magnitude (Voltage)"
        assert "[" not in rows[0][1]

    def test_manifest_may_still_record_measured_units(self):
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity="Voltage", unit="kV")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_WITH_PROVENANCE)

        manifest = _read_manifest(_unzip(result.content))
        assert manifest["column_measured_units"][1] == "kV"

    def test_data_only_export_needs_no_manifest_to_carry_the_unit(self):
        prep = PreparationSessionRegistry()
        sid = self._labeled_source(prep, quantity="ROCOF", unit="Hz/s")

        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)

        with pytest.raises(zipfile.BadZipFile):
            _unzip(result.content)
        rows = list(csv.reader(io.StringIO(result.content.decode("utf-8"))))
        assert rows[0][1] == "CBDK_V1 Magnitude (ROCOF) [Hz/s]"


class TestMeasuredUnitReUploadRoundTrip:
    """Measured Unit enhancement (DEC-080), task sections S/U/AS: re-
    uploading a cleaned export restores BOTH Engineering Quantity and
    Measured Unit deterministically from the header suffix, and
    exporting again is stable (never a duplicated suffix)."""

    def _first_export_sid(self, prep) -> str:
        content = b"Time,CBDK_V1 Magnitude\n2026-08-31 13:00:00,1.0\n2026-08-31 13:00:01,2.0\n"
        sid = _add_csv(prep, content)
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)
        set_column_engineering_quantity(
            workspace_id=WS, source_id=sid, column_index=1, engineering_quantity="Voltage", registry=prep,
        )
        set_column_measured_unit(workspace_id=WS, source_id=sid, column_index=1, measured_unit="kV", registry=prep)
        _confirm_absolute(prep, sid, column_index=0)
        return sid

    def test_full_round_trip_restores_and_re_exports_identically(self):
        prep = PreparationSessionRegistry()
        first_sid = self._first_export_sid(prep)
        first_result = export_preparation_source(
            workspace_id=WS, source_id=first_sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY,
        )
        assert first_result.content.decode("utf-8").splitlines()[0] == "Time,CBDK_V1 Magnitude (Voltage) [kV]"

        second_sid = _add_csv(prep, first_result.content, filename="re-uploaded.csv")
        set_header_row(workspace_id=WS, source_id=second_sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=second_sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, second_sid, 0)
        _mark_waveform(prep, second_sid, 1)  # restoration fires here
        _confirm_absolute(prep, second_sid, column_index=0)

        session = prep.get(WS, second_sid)
        assert session.working_overlay.column_engineering_quantities[column_key(None, 1)] == "Voltage"
        assert session.working_overlay.column_measured_units[column_key(None, 1)] == "kV"

        second_result = export_preparation_source(
            workspace_id=WS, source_id=second_sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY,
        )
        second_header = second_result.content.decode("utf-8").splitlines()[0]
        # Stable, never duplicated -- "... (Voltage) [kV] (Voltage) [kV]" would be a bug.
        assert second_header == "Time,CBDK_V1 Magnitude (Voltage) [kV]"

    def test_quantity_only_export_remains_backward_compatible_on_reupload(self):
        """An existing DEC-077-only export (no unit suffix) still
        restores the quantity alone (task section T)."""
        prep = PreparationSessionRegistry()
        content = b"Time,CBDK_V1 Magnitude (Voltage)\n2026-08-31 13:00:00,1.0\n2026-08-31 13:00:01,2.0\n"
        sid = _add_csv(prep, content, filename="legacy-export.csv")
        set_header_row(workspace_id=WS, source_id=sid, row_number=1, registry=prep)
        set_data_region(workspace_id=WS, source_id=sid, start_row=2, end_row=3, registry=prep)
        _mark_time_axis(prep, sid, 0)
        _mark_waveform(prep, sid, 1)  # restoration fires here

        session = prep.get(WS, sid)
        assert session.working_overlay.column_engineering_quantities[column_key(None, 1)] == "Voltage"
        assert column_key(None, 1) not in session.working_overlay.column_measured_units

        _confirm_absolute(prep, sid, column_index=0)
        result = export_preparation_source(workspace_id=WS, source_id=sid, registry=prep, mode=EXPORT_MODE_DATA_ONLY)
        assert result.content.decode("utf-8").splitlines()[0] == "Time,CBDK_V1 Magnitude (Voltage)"
