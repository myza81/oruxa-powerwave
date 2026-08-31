"""CSV/Excel preparation-source import orchestration (Slices 1-2, DEC-072).

Owns the upload -> validate -> accept-as-raw -> registry lifecycle for a
CSV or Excel preparation source. Deliberately the narrowest possible
slice:

    External CSV/Excel
        |
    Temporary immutable raw source   <- this module's own job, nothing more
        |
    Preparation Session (+ worksheet discovery for Excel)

Neither format's importer parses tabular structure at all -- no header
detection, no delimiter/column/time-axis inspection, no
`DisturbanceRecord` construction. "Safe acceptance" means exactly: a
real filename with the right extension, a non-empty upload, a size
within the configured limit, and -- for Excel only -- that the workbook
actually opens and has at least one worksheet. Both importers reuse the
exact same generic helpers (`app.services.upload_utils`) rather than
duplicating validation.

Never writes to StorageBackend or any persistent location -- the raw
bytes are held only in the caller-supplied
`PreparationSessionRegistry`'s in-memory store (see that module's own
docstring for why, and for the DEC-015/DEC-072 boundary this respects).
"""

from __future__ import annotations

import io
import uuid

from fastapi import UploadFile
from openpyxl import load_workbook

from app.domain.preparation_session import (
    FORMAT_CSV,
    FORMAT_EXCEL,
    STATUS_NEEDS_PREPARATION,
    PreparationSession,
    PreparationSessionSummary,
    WorksheetInfo,
)
from app.domain.source import utc_now
from app.services.errors import (
    EmptyWorkbookError,
    InvalidFileError,
    InvalidWorksheetIndexError,
    SourceNotFoundError,
    UploadTooLargeError,
    WorkbookParseError,
    WorksheetSelectionNotApplicableError,
)
from app.services.preparation_session_registry import PreparationSessionRegistry
from app.services.upload_utils import read_bounded, validate_suffix

_CSV_SUFFIXES = {".csv"}
#: .xlsx only -- see app.domain.preparation_session's own module
#: docstring for why legacy .xls is deliberately not supported this
#: slice (a separate, unmaintained xlrd dependency, not currently
#: justified).
_EXCEL_SUFFIXES = {".xlsx"}


async def import_csv_preparation_source(
    *,
    workspace_id: str,
    csv_upload: UploadFile,
    max_total_bytes: int,
    registry: PreparationSessionRegistry,
) -> PreparationSessionSummary:
    """Validate and accept one CSV file as raw, immutable preparation input.

    Raises app.services.errors.ImportServiceError subclasses on any
    failure (unsupported_file_type / invalid_file / upload_too_large --
    the exact same codes and HTTP mapping COMTRADE upload already uses;
    no new error taxonomy is introduced for Slice 1, per DEC-072's own
    "Readiness Issue model is Slice 6 scope, not Slice 1" guardrail).
    """
    filename = validate_suffix(csv_upload.filename, _CSV_SUFFIXES, "CSV")

    known_size = csv_upload.size or 0
    if known_size > max_total_bytes:
        raise UploadTooLargeError(
            f"Upload size ({known_size} bytes) exceeds the "
            f"{max_total_bytes // (1024 * 1024)} MB limit."
        )

    raw_bytes = await read_bounded(csv_upload, max_bytes=max_total_bytes, already_read=0)
    if not raw_bytes:
        raise InvalidFileError("CSV file is empty.")

    source_id = str(uuid.uuid4())
    summary = PreparationSessionSummary(
        source_id=source_id,
        workspace_id=workspace_id,
        original_filename=filename,
        source_format=FORMAT_CSV,
        original_byte_size=len(raw_bytes),
        status=STATUS_NEEDS_PREPARATION,
        created_at=utc_now(),
    )
    registry.add(PreparationSession(summary=summary, raw_bytes=raw_bytes))
    return summary


def _discover_worksheets(raw_bytes: bytes) -> list[WorksheetInfo]:
    """Open *raw_bytes* as an Excel workbook and return its worksheet
    structure only -- never cell values.

    `read_only=True` streams the workbook's XML parts directly rather
    than materializing every sheet's cell grid into memory (section:
    "do not fully materialize all worksheets just for discovery") --
    verified directly against openpyxl's own behavior for an in-memory
    `BytesIO` source: zero temporary files are created either way.
    `data_only=True` matters only if a later slice ever reads cell
    values through this same open (it does not here); harmless for pure
    structure discovery.

    `row_count`/`column_count` come from `Worksheet.max_row`/`max_column`,
    which openpyxl derives from the sheet XML's own `<dimension>` hint in
    read-only mode -- cheap, but not guaranteed present on every workbook
    (some writers omit it). Wrapped in its own `try/except` so a missing
    hint degrades to `None` rather than failing discovery entirely; this
    never triggers a full-sheet scan to compute an exact count.

    The workbook is always closed before returning (`finally`) -- no
    open workbook object or file handle is ever retained across a
    request; only the resulting plain `WorksheetInfo` list and the
    original `raw_bytes` are kept in the registry.

    Raises `WorkbookParseError` for anything that isn't a valid,
    openable .xlsx workbook (corrupt bytes, a non-Excel file renamed
    `.xlsx`, an unreadable internal structure) -- openpyxl raises a
    variety of exception types for these cases (`zipfile.BadZipFile`,
    `KeyError`, its own `InvalidFileException`, ...); all are treated
    identically here, mirroring how
    `app.services.import_service._classify_provider_error` collapses
    COMTRADE's own parser failures onto one user-safe error taxonomy
    rather than leaking an internal exception type to the client.
    Raises `EmptyWorkbookError` if the workbook opens but declares zero
    worksheets (structurally invalid as a preparation source).
    """
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise WorkbookParseError(f"Could not open the Excel workbook: {exc}") from exc

    try:
        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            raise EmptyWorkbookError("The workbook contains no worksheets.")

        worksheets: list[WorksheetInfo] = []
        for index, name in enumerate(sheet_names):
            worksheet = workbook[name]
            visible = getattr(worksheet, "sheet_state", "visible") == "visible"
            row_count: int | None = None
            column_count: int | None = None
            try:
                row_count = worksheet.max_row
                column_count = worksheet.max_column
            except Exception:
                # Best-effort only (see this function's own docstring) --
                # never fatal to discovery itself.
                pass
            worksheets.append(
                WorksheetInfo(
                    index=index, name=name, visible=visible,
                    row_count=row_count, column_count=column_count,
                )
            )
        return worksheets
    finally:
        workbook.close()


async def import_excel_preparation_source(
    *,
    workspace_id: str,
    excel_upload: UploadFile,
    max_total_bytes: int,
    registry: PreparationSessionRegistry,
) -> PreparationSessionSummary:
    """Validate, discover worksheets for, and accept one Excel workbook
    as raw, immutable preparation input (Slice 2).

    Sheets are never merged, concatenated, or otherwise combined (task's
    own "sheets remain independent" guardrail) -- `_discover_worksheets`
    only ever inspects structure, one workbook, in isolation.

    Auto-selection rule (task's own "Excel workbook with one sheet" /
    "may be automatically selected for convenience" guidance): when the
    workbook has EXACTLY one worksheet (regardless of its visible/hidden
    state), `selected_worksheet_index` starts at `0` -- deterministic,
    and still visibly reported to the caller as a real selection, never
    implying the workbook is otherwise "ready." A workbook with two or
    more worksheets (even if only one is visible) starts with
    `selected_worksheet_index=None`, requiring an explicit
    `PATCH .../preparation-sources/{id}` selection.

    Raises the same `ImportServiceError` taxonomy as CSV upload, plus
    `WorkbookParseError`/`EmptyWorkbookError` for a workbook that cannot
    be safely opened (see `_discover_worksheets`'s own docstring).
    """
    filename = validate_suffix(excel_upload.filename, _EXCEL_SUFFIXES, "Excel")

    known_size = excel_upload.size or 0
    if known_size > max_total_bytes:
        raise UploadTooLargeError(
            f"Upload size ({known_size} bytes) exceeds the "
            f"{max_total_bytes // (1024 * 1024)} MB limit."
        )

    raw_bytes = await read_bounded(excel_upload, max_bytes=max_total_bytes, already_read=0)
    if not raw_bytes:
        raise InvalidFileError("Excel file is empty.")

    worksheets = _discover_worksheets(raw_bytes)
    selected_worksheet_index = 0 if len(worksheets) == 1 else None

    source_id = str(uuid.uuid4())
    summary = PreparationSessionSummary(
        source_id=source_id,
        workspace_id=workspace_id,
        original_filename=filename,
        source_format=FORMAT_EXCEL,
        original_byte_size=len(raw_bytes),
        status=STATUS_NEEDS_PREPARATION,
        created_at=utc_now(),
        worksheets=tuple(worksheets),
        selected_worksheet_index=selected_worksheet_index,
    )
    registry.add(PreparationSession(summary=summary, raw_bytes=raw_bytes))
    return summary


def select_preparation_worksheet(
    *,
    workspace_id: str,
    source_id: str,
    worksheet_index: int,
    registry: PreparationSessionRegistry,
) -> PreparationSessionSummary:
    """Set the currently selected worksheet for an Excel preparation
    session (Slice 2).

    Stores ONLY the stable `index`/`name` identity already discovered at
    upload time -- never a header row, data region, or column mapping
    (those are later slices' own concepts; task's own explicit
    guardrail). Mutates the registry's own stored `PreparationSessionSummary`
    in place (the registry already holds this exact object by reference,
    so this is the single source of truth updated once, not a copy that
    could drift) and returns it.

    Raises `SourceNotFoundError` if no such preparation session exists,
    `WorksheetSelectionNotApplicableError` if it exists but has no
    worksheet concept at all (a CSV source), and
    `InvalidWorksheetIndexError` if `worksheet_index` is outside this
    workbook's own discovered range -- never silently clamped.
    """
    session = registry.get(workspace_id, source_id)
    if session is None:
        raise SourceNotFoundError(
            f"No preparation source '{source_id}' in workspace '{workspace_id}'."
        )

    worksheets = session.summary.worksheets
    if not worksheets:
        raise WorksheetSelectionNotApplicableError(
            "This preparation source has no worksheets to select (not an Excel workbook)."
        )

    if not isinstance(worksheet_index, int) or isinstance(worksheet_index, bool) or not (
        0 <= worksheet_index < len(worksheets)
    ):
        raise InvalidWorksheetIndexError(
            f"worksheet_index must be an integer in [0, {len(worksheets) - 1}]; got {worksheet_index!r}."
        )

    session.summary.selected_worksheet_index = worksheet_index
    return session.summary
