"""Cleaned Data Export (CSV/Excel ingestion Slice 12, DEC-072).

**Governing principle (task's own explicit framing): "Cleaned export =
the current Working Dataset as prepared by the engineer."** Not the
untouched raw source, not the canonical `DisturbanceRecord`, not a
silently repaired dataset -- exactly what the engineer's own header/
data-region/row-exclusion/column-role/cell-edit choices currently
produce, nothing more, nothing less.

**UAT enhancement (2026-09-04, DEC-074): now GATED on readiness.**
Originally available regardless of Powerwave readiness -- Slice 12's
own explicit design. Once cleaned export started serializing a
RESOLVED Time Axis (see below) that original policy stopped making
sense: there is no honest resolved Time column to build from an
unconfigured/unresolved/manual Time Axis. `export_preparation_source()`
now reuses `app.services.preparation_issue_service.build_issue_
summary()`'s own `is_ready` verdict as a real GATE (never a second,
narrower readiness policy of its own -- every current `blocking` issue
is already exactly a Time-Axis or Waveform-Channel finding, see
`app.services.readiness_service`'s own module docstring), plus the same
two additional capability constraints Slice 10's own canonical
conversion already enforces (`manual`/`unsupported` interpreter;
`sample_index` with no real interval) -- see `_ensure_exportable()`.
This module still recomputes readiness LIVE at export time rather than
trusting stale frontend state (task section V), and still records the
snapshot into the manifest exactly as before -- it is simply ALSO a
gate now, not merely informational.

**Row/column selection is a REUSE, not a re-derivation** -- the exact
same `app.services.preparation_preview_service.iterate_active_region_
rows()` single-pass streaming generator Slice 9's readiness validator
and Slice 10's own conversion service already use, filtered here to
non-excluded, non-header, in-active-region rows (identical filter to
Slice 10's own `convert_preparation_source()`). Column labels reuse
`preview_preparation_source()`'s own already-computed `column_labels`/
`column_roles` (which already encode the exact header-row-value /
neutral-spreadsheet-letter fallback task sections E/F ask for) --
this module adds only ONE new piece of logic on top: deduplicating
those labels for the columns actually being exported (task section G),
via the SAME stable-position `__{SpreadsheetLetter}` suffix strategy
Slice 10's own `_unique_channel_names()` established.

**UAT fix (2026-09-04, DEC-073, column roles): `not_assigned` columns
are omitted.** Owner-approved simplification of the column-role model
to exactly three roles (`not_assigned`/`time_axis`/`waveform` -- see
`app.domain.working_overlay`'s own module docstring for the full
rationale) makes `not_assigned` the single "not used by Powerwave"
state; this module omits every `not_assigned` column from the cleaned
table. The manifest's own `omitted_columns` entries record each
excluded column's `role` (always `not_assigned`) so the exact reason is
never left implicit -- a raw Time Axis source column is NOT one of
these entries (see the DEC-074 enhancement immediately below: it is
CONSUMED into the configured Time column, a different reason than
"not used by Powerwave," recorded separately under `exported_time.
source_columns`).

**UAT enhancement (2026-09-04, DEC-074): cleaned export now serializes
the RESOLVED/CONFIGURED Time Axis, not the original source Time Axis
columns.** Supersedes this module's own original Slice 12 policy
("Time columns are never touched" -- an original source Time Axis
column's own current working value, exported verbatim). Owner-approved
direction: re-uploading a cleaned file should require MINIMAL repeated
preparation, which means the export must carry the engineer's already-
resolved date-order/unit/interval/reconstruction choice forward, not
force it to be re-made on every re-upload.

The exported table is now exactly ONE standardized `Time`/`Time (s)`
column (see `_build_configured_time_column()` below) followed by every
Waveform-role column in source order -- the original source Time Axis
column(s) never appear in the cleaned table at all (their values are
CONSUMED to build the one configured Time column, not omitted the same
way a `not_assigned` column is; the manifest's own `exported_time.
source_columns` records which raw columns produced it). This makes a
usable, resolved Time Axis (plus at least one Waveform column) a
REQUIRED precondition for export now -- see `_ensure_exportable()`
below -- a real, intentional behavior change from the earlier "export
regardless of readiness" policy (task section B's own explicit
"supersedes the earlier policy" instruction).

**No new inference happens here** (task section L): every per-row Time
value comes from re-calling the ALREADY-CONFIRMED time-axis
interpreter's own `build_preview_rows()` -- the EXACT SAME Protocol
method `app.services.preparation_conversion_service`'s own canonical
`DisturbanceRecord` construction already calls -- over the FULL active
region, then reusing `app.services.time_axis_normalization`'s own
shared parse/canonicalize helpers (never a second, divergent
implementation; task section Z's own explicit "must agree" requirement).
This module's own job stays narrowly: format the already-resolved
native value into the ONE deterministic export representation (ISO-8601
for `FAMILY_ABSOLUTE`, fixed-precision relative seconds otherwise) --
never re-running date-order elimination, re-estimating cadence, or
inventing a timezone/date.

**Read-only, by construction**: this module calls no `app.domain.
working_overlay` mutation function anywhere, and captures/re-verifies
`WorkingOverlay.revision` around the whole export (task section W) --
if a concurrent mutation is somehow observed mid-export,
`ExportRevisionChangedError` discards the whole attempt rather than
mixing rows from two different working-overlay states, mirroring Slice
10's own `ConversionRevisionChangedError` precedent exactly.

**Performance** (task section Y): one single streaming pass over
`iterate_active_region_rows()` builds the export table's rows
incrementally -- for Excel, `openpyxl.Workbook(write_only=True)` +
`worksheet.append()` writes each row directly into the underlying
zip/XML stream rather than building an in-memory cell-object graph,
matching this module's own "never build raw + working + export full
copies simultaneously" requirement. CSV uses the stdlib `csv.writer`
over an `io.StringIO`, which is already effectively O(rows) with no
second full-file materialization beyond the one `iterate_active_region_
rows()` itself already performs.

**Packaging** (task section AB): a single ZIP bundle
(`<base>_cleaned.zip`) containing the cleaned CSV/XLSX plus a sidecar
`<base>_cleaned.manifest.json` -- one download action, no separate
packaging framework (`zipfile`, stdlib only).

**UAT enhancement (2026-09-04): manifest/provenance is now OPTIONAL,
not forced on every download.** Owner-approved problem: a normal
engineer only wants the reusable cleaned CSV/XLSX itself and should
never be handed a `.zip` (let alone have to understand a sidecar
`manifest.json`) merely to get it. `export_preparation_source()` now
takes an explicit `mode` -- `EXPORT_MODE_DATA_ONLY` (the new default:
returns the cleaned CSV/XLSX bytes directly, no ZIP, no manifest built
at all) or `EXPORT_MODE_WITH_PROVENANCE` (the original Slice 12/DEC-074
behavior, unchanged: cleaned CSV/XLSX + manifest inside one ZIP).
Provenance capability is NOT removed -- every manifest field, the
gating rules, and the configured-Time-Axis derivation are identical to
before; only the DEFAULT shape of what a plain "Export Cleaned Data"
click returns has changed. Both modes share the exact same gating
(`_ensure_exportable()`), the exact same configured-Time-Axis/waveform
row construction, and therefore always contain byte-identical cleaned
data for the same working-overlay revision -- `mode` only decides
whether that same artifact is handed back directly or bundled with a
manifest. The manifest is built and serialized ONLY for
`EXPORT_MODE_WITH_PROVENANCE` (task section M's own "don't build/
serialize the manifest unnecessarily for the default path" efficiency
note) -- `EXPORT_MODE_DATA_ONLY` never constructs `_build_manifest()`'s
own edited/cleared-cell provenance counts or JSON payload at all.
"""

from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook

from app.domain.channel_classification import (
    ENGINEERING_QUANTITY_UNDEFINED,
    encode_engineering_quantity_and_unit_suffix,
)
from app.domain.preparation_issue import SEVERITY_BLOCKING
from app.domain.preparation_session import FORMAT_CSV, PreparationSession
from app.domain.time_axis import (
    FAMILY_ABSOLUTE,
    FAMILY_PARTIAL,
    INTERPRETER_ID_MANUAL,
    INTERPRETER_ID_UNSUPPORTED,
    PROVENANCE_RECONSTRUCTED,
    TimeAxisSampleRow,
    is_time_axis_resolved,
)
from app.domain.working_overlay import OVERRIDE_KIND_CLEAR, OVERRIDE_KIND_EDIT, ROLE_NOT_ASSIGNED, ROLE_WAVEFORM
from app.services.errors import (
    ExportNotReadyError,
    ExportRequiresIntervalError,
    ExportRevisionChangedError,
    ExportTimeAxisValueError,
    ExportUnsupportedInterpreterError,
    SourceNotFoundError,
    WorksheetNotSelectedError,
)
from app.services.preparation_issue_service import build_issue_summary
from app.services.preparation_preview_service import (
    _spreadsheet_column_label,
    iterate_active_region_rows,
    preview_preparation_source,
)
from app.services.preparation_session_registry import PreparationSessionRegistry
from app.services.time_axis_normalization import (
    format_absolute_iso,
    format_relative_seconds,
    parse_native_time_value,
    relative_seconds,
    seconds_from_midnight,
)
from app.services.time_axis_service import get_time_axis_summary, resolve_interpreter

#: Task section T's own "avoid creating a huge manifest unnecessarily"
#: allowance -- listing up to this many excluded row numbers is cheap
#: and useful; beyond it, only the count is reported (plus a truncation
#: flag), never an unbounded list. Matches `app.domain.working_overlay.
#: MAX_OPERATION_HISTORY`'s own bound for the same "generous but not
#: unlimited" reasoning.
MAX_MANIFEST_EXCLUDED_ROWS_LISTED = 200

#: Excel worksheet name constraints (task section C): at most 31
#: characters, and none of `: \ / ? * [ ]` -- Excel's own real
#: constraints, not an invented rule. A name that becomes empty after
#: sanitization (e.g. the source name was entirely invalid characters)
#: falls back to this deterministic default, matching a normal new
#: workbook's own first-sheet name.
_EXCEL_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
_EXCEL_MAX_SHEET_NAME_LENGTH = 31
_DEFAULT_SHEET_NAME = "Sheet1"

#: Filename sanitization (task section AC): keep only characters safe
#: in a downloaded filename across common filesystems -- alphanumerics,
#: dash, underscore, and space (collapsed to underscore); everything
#: else is dropped, never trusted from the raw uploaded filename
#: (which may contain path separators or other unsafe characters).
_UNSAFE_FILENAME_CHARS = re.compile(r"[^A-Za-z0-9._ -]")
_MAX_BASE_FILENAME_LENGTH = 100
_DEFAULT_BASE_FILENAME = "recording"

MANIFEST_VERSION = 1

#: `export_preparation_source()`'s own `mode` values (2026-09-04 UAT
#: enhancement, above). `EXPORT_MODE_DATA_ONLY` is the new default --
#: matches the owner-approved "give me the cleaned reusable file, not a
#: ZIP" default action; `EXPORT_MODE_WITH_PROVENANCE` is the unchanged
#: Slice 12/DEC-074 ZIP+manifest bundle, kept as an explicit opt-in.
EXPORT_MODE_DATA_ONLY = "data_only"
EXPORT_MODE_WITH_PROVENANCE = "with_provenance"

_CSV_CONTENT_TYPE = "text/csv"
_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(slots=True)
class ExportResult:
    """The finished, ready-to-return export artifact. For
    `EXPORT_MODE_DATA_ONLY`, `content` is the cleaned CSV/XLSX bytes
    directly (`media_type` is the real CSV/XLSX content type); for
    `EXPORT_MODE_WITH_PROVENANCE`, `content` is a ZIP bundle containing
    the cleaned CSV/XLSX plus its own manifest JSON
    (`media_type="application/zip"`). Never a filesystem path (task
    section Z: no durable storage is introduced by this module) -- the
    caller (the API route) returns `content` directly as an HTTP
    response body."""

    filename: str
    content: bytes
    media_type: str = "application/zip"


def _resolve_session(*, workspace_id: str, source_id: str, registry: PreparationSessionRegistry) -> PreparationSession:
    session = registry.get(workspace_id, source_id)
    if session is None:
        raise SourceNotFoundError(f"No preparation source '{source_id}' in workspace '{workspace_id}'.")
    return session


def _resolve_worksheet_index(session: PreparationSession) -> int | None:
    worksheets = session.summary.worksheets
    if not worksheets:
        return None
    if session.summary.selected_worksheet_index is None:
        raise WorksheetNotSelectedError(
            "This workbook has more than one worksheet; select one with "
            "PATCH .../preparation-sources/{source_id} before exporting it."
        )
    return session.summary.selected_worksheet_index


def _sanitize_base_filename(original_filename: str) -> str:
    """`event.csv` -> `event` (task section AC's own worked example) --
    strips the original extension, keeps only filesystem-safe
    characters, collapses runs of whitespace to a single underscore,
    and falls back to a deterministic default if nothing safe survives
    (an entirely non-ASCII or entirely-symbolic original filename)."""
    stem = original_filename.rsplit(".", 1)[0] if "." in original_filename else original_filename
    cleaned = _UNSAFE_FILENAME_CHARS.sub("", stem).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = cleaned[:_MAX_BASE_FILENAME_LENGTH].strip("._")
    return cleaned or _DEFAULT_BASE_FILENAME


def _sanitize_sheet_name(name: str) -> str:
    """Task section C's own explicit rule: preserve the original
    selected worksheet name when practical; if it becomes invalid
    (empty, or over Excel's own 31-character limit, after stripping
    `: \\ / ? * [ ]`) apply the deterministic `Sheet1` fallback rather
    than guessing at a "close enough" alternative."""
    cleaned = _EXCEL_INVALID_SHEET_CHARS.sub("", name).strip()
    cleaned = cleaned[:_EXCEL_MAX_SHEET_NAME_LENGTH]
    return cleaned or _DEFAULT_SHEET_NAME


def _unique_export_names(column_labels: list[str], included_column_indices: list[int]) -> dict[int, str]:
    """One deterministic, unique export column name per INCLUDED column
    (task section G) -- the same strategy Slice 10's own canonical
    conversion `_unique_channel_names()` established: the FIRST included
    column to use a given label keeps it verbatim; every LATER included
    column sharing that same label gets a stable `__{SpreadsheetLetter}`
    suffix (`Voltage`, `Voltage__C`, `Voltage__D`). Deliberately scoped
    to only the columns actually being exported -- an omitted
    `not_assigned` column's own label never consumes a disambiguation
    slot for columns that ARE exported. `column_labels` already carries
    `preview_preparation_source()`'s own header-row-value / neutral-
    fallback logic (task sections E/F); this function's only job is
    uniqueness on top of that, never re-deriving the label itself."""
    seen: dict[str, int] = {}
    result: dict[int, str] = {}
    for column_index in included_column_indices:
        label = (
            column_labels[column_index]
            if column_index < len(column_labels)
            else _spreadsheet_column_label(column_index)
        )
        count = seen.get(label, 0)
        seen[label] = count + 1
        result[column_index] = label if count == 0 else f"{label}__{_spreadsheet_column_label(column_index)}"
    return result


def _cell_export_value(value: Any) -> Any:
    """Task section L: a genuinely empty cell (raw `None`, or a working
    CLEAR override -- both already surface as `None` in `PreviewRow.
    cells`) exports as an empty field/cell, never the literal text
    `"None"`/`"NaN"`/`"null"`/`0`. Every other value passes through
    UNCHANGED -- this function never coerces, parses, or reinterprets a
    non-empty value (task's own explicit "do not coerce ambiguous
    non-empty text" guardrail)."""
    return "" if value is None else value


def _ensure_exportable(*, issue_summary, time_axis_summary) -> None:
    """Task section B: a reusable cleaned export now requires a usable,
    resolved Time Axis plus at least one Waveform column. Reuses
    `is_ready` directly as the primary gate (never a second, narrower
    readiness policy of its own -- see this module's own docstring for
    why that verdict already covers exactly Time-Axis/Waveform-Channel
    blocking conditions), plus `app.domain.time_axis.
    is_time_axis_resolved()` -- the SAME shared eligibility check the
    Data Preview's own "Configured Time" column now reuses too (a
    2026-09-04, DEC-075 enhancement) -- to decide WHETHER a standardized
    Time column can honestly be built at all. The two specific checks
    below stay separate only to pick the right, specific error message;
    by the time either is reached, `is_ready` has already ruled out
    every OTHER way `is_time_axis_resolved()` could be `False` (an
    unconfigured/unresolved/stale-role-reference Time Axis is always
    itself a BLOCKING readiness issue -- see
    `app.services.readiness_service`'s own module docstring)."""
    if not issue_summary.is_ready:
        blocking_messages = [i.message for i in issue_summary.issues if i.severity == SEVERITY_BLOCKING]
        raise ExportNotReadyError(
            "This source is not yet ready for a reusable cleaned export: " + " ".join(blocking_messages)
        )
    if is_time_axis_resolved(time_axis_summary):
        return
    if time_axis_summary.interpreter_id in (INTERPRETER_ID_MANUAL, INTERPRETER_ID_UNSUPPORTED):
        raise ExportUnsupportedInterpreterError(
            "The active Time Axis configuration does not parse real per-row values from this source's own "
            "columns -- assign a real interpreter (Absolute Datetime, Date + Time, Elapsed Time, Sample Index, "
            "or Repeated Timestamp) before exporting a reusable cleaned file."
        )
    raise ExportRequiresIntervalError(
        "A sampling interval or sampling rate is required before a reusable cleaned file can be exported. "
        "Return to Time Axis configuration and provide one."
    )


@dataclass(slots=True)
class _ConfiguredTimeColumn:
    """The one standardized Time column this export now produces (task
    sections C-N), plus everything the manifest's own `exported_time`
    section needs to describe it (task sections T-V) -- computed
    together, in the SAME pass, so the manifest can never drift from
    what was actually written into the exported table."""

    column_name: str
    values: list[str]
    export_representation: str
    timezone_present: bool
    source_offset_seconds: float | None


def _build_configured_time_column(
    *, interpreter, time_axis_samples: list[TimeAxisSampleRow], time_axis_summary,
) -> _ConfiguredTimeColumn:
    """Re-calls the ALREADY-CONFIRMED interpreter's own
    `build_preview_rows()` over the FULL active region (never a bounded
    sample, never a second reconstruction -- the exact same call
    `app.services.preparation_conversion_service` already makes for
    canonical `DisturbanceRecord` construction), then formats each
    row's already-resolved native value via `app.services.
    time_axis_normalization`'s own shared helpers: absolute timestamps
    stay real per-row ISO-8601 values (never collapsed to relative
    seconds -- task section D's own explicit "one standardized ISO-
    8601-style timestamp column" requirement), every other family
    becomes fixed-precision seconds relative to the first active row
    (task section F/G/I). One value per `time_axis_samples` entry, in
    the SAME order (task section M's own one-to-one row-alignment
    requirement) -- this function never sorts, drops, or inserts a row.
    """
    family = time_axis_summary.family
    preview_rows = interpreter.build_preview_rows(
        samples=time_axis_samples, resolved_options=time_axis_summary.options,
        resolved_unit=time_axis_summary.unit, resolved_interval_seconds=time_axis_summary.interval_seconds,
        limit=len(time_axis_samples),
    )
    natives: list[Any] = []
    for row in preview_rows:
        if row.interpreted is None:
            raise ExportTimeAxisValueError(
                f"Row {row.row_number}'s Time Axis value could not be interpreted under the confirmed configuration."
            )
        native = parse_native_time_value(row.interpreted, family=family)
        if native is None:
            raise ExportTimeAxisValueError(
                f"Row {row.row_number}'s Time Axis value '{row.interpreted}' could not be parsed during export."
            )
        natives.append(native)

    if family == FAMILY_ABSOLUTE:
        timezone_present = bool(natives) and natives[0].tzinfo is not None
        return _ConfiguredTimeColumn(
            column_name="Time", values=[format_absolute_iso(n) for n in natives],
            export_representation="iso8601", timezone_present=timezone_present, source_offset_seconds=None,
        )

    try:
        rel = relative_seconds(natives, family=family)
    except TypeError as exc:
        raise ExportTimeAxisValueError(
            "A row mixes a timezone-aware timestamp with the first active row's own naive (or vice-versa) "
            "timestamp -- cannot compute relative time."
        ) from exc
    if not natives:
        source_offset_seconds = None
    elif family == FAMILY_PARTIAL:
        source_offset_seconds = seconds_from_midnight(natives[0])
    else:
        source_offset_seconds = float(natives[0])
    return _ConfiguredTimeColumn(
        column_name="Time (s)", values=[format_relative_seconds(v) for v in rel],
        export_representation="relative_seconds", timezone_present=False,
        source_offset_seconds=source_offset_seconds,
    )


def export_preparation_source(
    *, workspace_id: str, source_id: str, registry: PreparationSessionRegistry,
    mode: str = EXPORT_MODE_DATA_ONLY,
) -> ExportResult:
    """Export the CURRENT Working Dataset of one CSV/Excel preparation
    source. Requires a usable, resolved Time Axis plus at least one
    Waveform column (see `_ensure_exportable()` -- a 2026-09-04,
    DEC-074 change from the earlier "available regardless of readiness"
    policy, once export started serializing a RESOLVED Time Axis rather
    than the raw source columns verbatim). Raises an
    `ImportServiceError` subclass (never a `PreparationIssue`) for every
    runtime failure; never mutates the preparation session, the working
    overlay, or the raw source in any way, in either mode.

    `mode=EXPORT_MODE_DATA_ONLY` (the default): returns the cleaned
    CSV/XLSX bytes directly -- no ZIP, no manifest built at all (task
    section M). `mode=EXPORT_MODE_WITH_PROVENANCE`: returns the same
    cleaned CSV/XLSX bundled with a sidecar provenance manifest inside
    one ZIP -- the original Slice 12/DEC-074 behavior, unchanged. Both
    modes share identical gating and identical cleaned-data construction
    (task section J's own "must agree" requirement) -- only the return
    shape differs.
    """
    if mode not in (EXPORT_MODE_DATA_ONLY, EXPORT_MODE_WITH_PROVENANCE):
        raise ValueError(f"Unknown export mode: {mode!r}")
    session = _resolve_session(workspace_id=workspace_id, source_id=source_id, registry=registry)
    worksheet_index = _resolve_worksheet_index(session)
    captured_revision = session.working_overlay.revision

    # Readiness is captured LIVE, exactly like Slice 10's own conversion
    # -- never trusting stale frontend state (task section V) -- and is
    # now ALSO the primary export gate (task section B), not merely a
    # manifest snapshot.
    issue_summary = build_issue_summary(workspace_id=workspace_id, source_id=source_id, registry=registry)
    time_axis_summary = get_time_axis_summary(workspace_id=workspace_id, source_id=source_id, registry=registry)
    _ensure_exportable(issue_summary=issue_summary, time_axis_summary=time_axis_summary)
    interpreter = resolve_interpreter(
        column_count=len(time_axis_summary.column_indices), requested_interpreter_id=time_axis_summary.interpreter_id,
    )

    preview = preview_preparation_source(workspace_id=workspace_id, source_id=source_id, offset=0, limit=1, registry=registry)
    column_labels = preview.column_labels
    column_roles = preview.column_roles
    column_engineering_quantities = preview.column_engineering_quantities
    column_measured_units = preview.column_measured_units
    # DEC-073: cleaned export omits every `not_assigned` column (role
    # recorded on each entry so the manifest states WHY). The raw Time
    # Axis source column(s) are NOT one of these entries -- see
    # `exported_time.source_columns` in the manifest below for how they
    # are accounted for instead (task section G/V).
    waveform_column_indices = sorted(c for c, role in enumerate(column_roles) if role == ROLE_WAVEFORM)
    omitted_columns = [
        {"column_index": c, "label": column_labels[c] if c < len(column_labels) else _spreadsheet_column_label(c), "role": role}
        for c, role in enumerate(column_roles)
        if role == ROLE_NOT_ASSIGNED
    ]
    # Engineering Quantity + Measured Unit enhancement (DEC-077/DEC-080,
    # task section N/O/P): each Waveform column's own label gets a
    # strict, self-describing `" (<Engineering Quantity>) [<Measured
    # Unit>]"` suffix BEFORE uniqueness dedup runs (so two identically-
    # labeled-but-differently-classified columns can naturally
    # disambiguate via their own suffix first) -- an Undefined quantity
    # leaves the label unchanged (never a noisy `"(Undefined)"`), and a
    # blank unit omits the bracket entirely (never a noisy `"[]"`, task
    # section Q). Reuses the SAME encode function the re-upload suffix
    # parser is the exact inverse of -- never two independently-
    # maintained grammars.
    suffixed_labels = list(column_labels)
    for c in waveform_column_indices:
        quantity = (
            column_engineering_quantities[c] if c < len(column_engineering_quantities) else ENGINEERING_QUANTITY_UNDEFINED
        )
        unit = column_measured_units[c] if c < len(column_measured_units) else ""
        label = suffixed_labels[c] if c < len(suffixed_labels) else _spreadsheet_column_label(c)
        suffixed_labels[c] = encode_engineering_quantity_and_unit_suffix(label, quantity, unit)
    export_name_by_column = _unique_export_names(suffixed_labels, waveform_column_indices)
    # DEC-074, task section R: the configured Time column is always
    # FIRST, regardless of its own source column's original position --
    # a deliberate, documented exception to "preserve absolute source
    # column order" (which still governs the WAVEFORM columns among
    # themselves, immediately after it). Assembled below once the time
    # column's own name/values are known.

    # Single streaming pass (task section Y): builds the time-axis
    # sample list and the waveform cell-value rows TOGETHER, in
    # lockstep, so they stay row-aligned one-to-one without a second
    # pass (task section M).
    time_axis_samples: list[TimeAxisSampleRow] = []
    waveform_rows: list[list[Any]] = []
    excluded_row_numbers: list[int] = []
    for row in iterate_active_region_rows(session, worksheet_index=worksheet_index):
        if row.excluded:
            excluded_row_numbers.append(row.row_number)
        # Task section E: the header row is schema, never an exported
        # data row, even if the active data region happens to overlap
        # it (a deliberate, documented choice for that overlap case).
        if row.is_header or not row.in_active_region or row.excluded:
            continue
        values = tuple(row.cells[c] if c < len(row.cells) else None for c in time_axis_summary.column_indices)
        time_axis_samples.append(TimeAxisSampleRow(row_number=row.row_number, values=values))
        waveform_rows.append([_cell_export_value(row.cells[c] if c < len(row.cells) else None) for c in waveform_column_indices])

    configured_time = _build_configured_time_column(
        interpreter=interpreter, time_axis_samples=time_axis_samples, time_axis_summary=time_axis_summary,
    )
    if len(configured_time.values) != len(waveform_rows):
        raise ExportTimeAxisValueError("Configured Time values did not align one-to-one with exported waveform rows.")

    header_names = [configured_time.column_name] + [export_name_by_column[c] for c in waveform_column_indices]
    exported_rows = [[time_value] + waveform_row for time_value, waveform_row in zip(configured_time.values, waveform_rows)]

    # Edit/clear provenance counts (task section U) -- one cheap pass
    # over just this worksheet's own sparse overrides (proportional to
    # edit count, never to dataset size), scoped the SAME way Slice 10's
    # own `excluded_row_count` provenance field already is.
    edited_cell_count = 0
    cleared_cell_count = 0
    for (ws, _row_number, _column_index), override in session.working_overlay.cell_overrides.items():
        if ws != worksheet_index:
            continue
        if override.kind == OVERRIDE_KIND_EDIT:
            edited_cell_count += 1
        elif override.kind == OVERRIDE_KIND_CLEAR:
            cleared_cell_count += 1

    region = session.working_overlay.data_region.get(worksheet_index)
    base_name = _sanitize_base_filename(session.summary.original_filename)

    if session.summary.source_format == FORMAT_CSV:
        artifact_bytes = _build_csv_artifact(header_names, exported_rows)
        artifact_filename = f"{base_name}_cleaned.csv"
        artifact_content_type = _CSV_CONTENT_TYPE
    else:
        sheet_name = _sanitize_sheet_name(session.summary.worksheets[worksheet_index].name)
        artifact_bytes = _build_xlsx_artifact(sheet_name, header_names, exported_rows)
        artifact_filename = f"{base_name}_cleaned.xlsx"
        artifact_content_type = _XLSX_CONTENT_TYPE

    # `EXPORT_MODE_DATA_ONLY` (the default): the manifest is never built
    # or serialized at all (task section M's own "don't build/serialize
    # the manifest unnecessarily for the default path" efficiency note)
    # -- only the revision race check below still applies, identically
    # to `EXPORT_MODE_WITH_PROVENANCE`.
    if mode == EXPORT_MODE_WITH_PROVENANCE:
        manifest = _build_manifest(
            session=session, worksheet_index=worksheet_index, captured_revision=captured_revision,
            issue_summary=issue_summary, time_axis_summary=time_axis_summary,
            omitted_columns=omitted_columns, column_roles=column_roles,
            column_engineering_quantities=column_engineering_quantities,
            column_measured_units=column_measured_units,
            edited_cell_count=edited_cell_count, cleared_cell_count=cleared_cell_count,
            excluded_row_numbers=excluded_row_numbers, exported_row_count=len(exported_rows),
            artifact_filename=artifact_filename, configured_time=configured_time, column_labels=column_labels,
        )
        manifest_filename = f"{base_name}_cleaned.manifest.json"
        manifest_bytes = json.dumps(manifest, indent=2, sort_keys=False).encode("utf-8")

    # Revision race protection (task section W) -- verify nothing
    # mutated the working overlay while this export was being built,
    # mirroring Slice 10's own `ConversionRevisionChangedError`
    # precedent exactly. Export never registers/persists anything, so
    # there is no "half from one revision" state to leave behind either
    # way -- this simply refuses to hand back a bundle that may have
    # mixed two different working-overlay states. Applies identically to
    # both modes.
    if session.working_overlay.revision != captured_revision:
        raise ExportRevisionChangedError(
            "This preparation source changed while the export was being built -- please retry."
        )

    if mode != EXPORT_MODE_WITH_PROVENANCE:
        return ExportResult(filename=artifact_filename, content=artifact_bytes, media_type=artifact_content_type)

    zip_bytes = _build_zip(artifact_filename, artifact_bytes, manifest_filename, manifest_bytes)
    return ExportResult(filename=f"{base_name}_cleaned.zip", content=zip_bytes)


def _build_csv_artifact(header_names: list[str], rows: list[list[Any]]) -> bytes:
    """Task section AK: a normalized, deterministic CSV dialect --
    comma delimiter, UTF-8, the stdlib `csv` module's own standard
    newline handling (`\\r\\n` via `writer` default) -- never an attempt
    to preserve whatever delimiter/quoting the ORIGINAL source file
    happened to use."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(header_names)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def _build_xlsx_artifact(sheet_name: str, header_names: list[str], rows: list[list[Any]]) -> bytes:
    """Task section AI: a single CLEAN tabular worksheet -- no attempt
    to preserve the original workbook's styling, formulas, charts,
    macros, or merged cells, and no recalculation of any formula (task
    section AJ: the WORKING cell value, exactly as the preparation
    overlay already represents it, is written verbatim). `write_only`
    mode (task section Y) streams each row directly into the underlying
    zip/XML rather than building an in-memory cell-object graph for the
    whole sheet."""
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(header_names)
    for row in rows:
        worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _exported_time_manifest(
    *, configured_time: _ConfiguredTimeColumn, time_axis_summary, column_labels: list[str],
) -> dict[str, Any]:
    """Task sections T-V: the manifest's own dedicated provenance for
    the ONE standardized Time column this export now writes -- records
    which raw source column(s) it was CONSUMED from (`source_columns`,
    by stable index and label -- task section V's own explicit
    traceability requirement; the raw source itself still holds the
    actual original values, untouched), plus the same
    family/provenance/interpreter/date-order/interval facts the rest of
    this manifest already carries at the top level, gathered into one
    self-contained section so a reader never has to cross-reference
    other manifest keys to understand this one column."""
    source_columns = [
        {
            "column_index": c,
            "label": column_labels[c] if c < len(column_labels) else _spreadsheet_column_label(c),
        }
        for c in time_axis_summary.column_indices
    ]
    return {
        "column_name": configured_time.column_name,
        "source_columns": source_columns,
        "family": time_axis_summary.family,
        "provenance": time_axis_summary.provenance,
        "interpreter_id": time_axis_summary.interpreter_id,
        "date_order": (time_axis_summary.options or {}).get("date_order"),
        "interval_seconds": time_axis_summary.interval_seconds,
        "export_representation": configured_time.export_representation,
        "timezone_present": configured_time.timezone_present,
        "source_offset_seconds": configured_time.source_offset_seconds,
        "reconstructed": time_axis_summary.provenance == PROVENANCE_RECONSTRUCTED,
    }


def _build_manifest(
    *, session: PreparationSession, worksheet_index: int | None, captured_revision: int,
    issue_summary, time_axis_summary, omitted_columns: list[dict], column_roles: list[str],
    column_engineering_quantities: list[str], column_measured_units: list[str],
    edited_cell_count: int, cleared_cell_count: int, excluded_row_numbers: list[int],
    exported_row_count: int, artifact_filename: str,
    configured_time: _ConfiguredTimeColumn, column_labels: list[str],
) -> dict[str, Any]:
    """Task section S: stable, machine-readable manifest keys only --
    never a raw Python object repr, never an attempt to recreate the
    original raw source (task section R: the raw bytes remain
    separately immutable in the `PreparationSession` itself)."""
    excluded_row_numbers = sorted(excluded_row_numbers)
    truncated = len(excluded_row_numbers) > MAX_MANIFEST_EXCLUDED_ROWS_LISTED
    listed_excluded_rows = excluded_row_numbers[:MAX_MANIFEST_EXCLUDED_ROWS_LISTED]

    worksheet_name = (
        session.summary.worksheets[worksheet_index].name if worksheet_index is not None else None
    )
    region = session.working_overlay.data_region.get(worksheet_index)

    return {
        "manifest_version": MANIFEST_VERSION,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "exported_file": artifact_filename,
        "source_format": session.summary.source_format,
        "original_filename": session.summary.original_filename,
        "worksheet_name": worksheet_name,
        "worksheet_index": worksheet_index,
        "preparation_revision": captured_revision,
        "header_row": session.working_overlay.header_row.get(worksheet_index),
        "data_region": (
            {"start_row": region.start_row, "end_mode": region.end_mode, "end_row": region.end_row}
            if region is not None else None
        ),
        "exported_row_count": exported_row_count,
        "excluded_row_count": len(excluded_row_numbers),
        "excluded_rows": listed_excluded_rows,
        "excluded_rows_truncated": truncated,
        "omitted_columns": omitted_columns,
        "column_roles": column_roles,
        # Engineering Quantity + Measured Unit enhancement (DEC-077/
        # DEC-080, task section V/Y): the optional manifest MAY continue
        # recording these, but re-upload restoration never depends on it
        # -- the cleaned CSV/XLSX header itself already carries the same
        # information via its own strict suffix grammar (see
        # encode_engineering_quantity_and_unit_suffix()).
        "column_engineering_quantities": column_engineering_quantities,
        "column_measured_units": column_measured_units,
        "edited_cell_count": edited_cell_count,
        "cleared_cell_count": cleared_cell_count,
        "time_family": time_axis_summary.family,
        "time_provenance": time_axis_summary.provenance,
        "interpreter_id": time_axis_summary.interpreter_id,
        "time_unit": time_axis_summary.unit,
        "time_interval_seconds": time_axis_summary.interval_seconds,
        "reconstructed_timing": time_axis_summary.provenance == PROVENANCE_RECONSTRUCTED,
        "exported_time": _exported_time_manifest(
            configured_time=configured_time, time_axis_summary=time_axis_summary, column_labels=column_labels,
        ),
        "readiness": {
            "is_ready": issue_summary.is_ready,
            "blocking_count": issue_summary.blocking_count,
            "warning_count": issue_summary.warning_count,
            "info_count": issue_summary.info_count,
        },
    }


def _build_zip(artifact_filename: str, artifact_bytes: bytes, manifest_filename: str, manifest_bytes: bytes) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(artifact_filename, artifact_bytes)
        archive.writestr(manifest_filename, manifest_bytes)
    return buffer.getvalue()
