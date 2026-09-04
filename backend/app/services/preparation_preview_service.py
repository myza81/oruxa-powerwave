"""Paged raw-data preview for CSV/Excel preparation sources (Slices 3-5, DEC-072).

Owns exactly one job: given an already-accepted `PreparationSession`
(Slices 1-2), return a bounded WINDOW of its rows/cells -- never the
whole dataset, never a type coercion beyond what the underlying reader
naturally exposes. This is strictly an inspection surface:

    PreparationSession (raw, immutable)
            |
    preview_preparation_source() (this module)
            |
    raw page  +  WorkingOverlay (Slice 4 edits/exclusions, Slice 5
    header/data-region/column-role state), applied at read time
            |
    a bounded page of WORKING rows -- never cached beyond one request,
    never mutates the session's own raw_bytes OR its working_overlay

No `DisturbanceRecord` is read or produced here. Nothing in this module
interprets timestamps, values, or engineering meaning -- see this
feature's own explicit non-goals in
docs/project-memory/CSV_EXCEL_INGESTION_ARCHITECTURE.md. Slice 5's own
header/column-role state changes WHICH ROW SUPPLIES LABELS and WHAT A
COLUMN IS CALLED, never what a cell's VALUE means.

CSV strategy (task's own "avoid parsing the entire CSV into a full
DataFrame just to return 200 rows"): the raw bytes are decoded once per
request and streamed through `csv.reader` -- never `pandas.read_csv`.
Because the in-memory bytes have no index, reaching row N still means
iterating from row 0 (task's own "acceptable initially if documented and
bounded" allowance) -- but the exact row/column TOTALS only need to be
computed once per session: `PreparationSession.cached_row_count`/
`cached_column_count` (see that dataclass's own docstring) are memoized
on the first preview request (any page) via `ensure_csv_totals_cached()`
-- Slice 4/5's own coordinate-bounds validation
(`app.services.working_overlay_service`) reuses this exact same function
rather than a second scan implementation.

Excel strategy: reuses `openpyxl`'s `read_only=True` streaming mode
(same choice as Slice 2's own worksheet discovery), reopening the
workbook fresh per request (never held open across requests -- see
Slice 2's own precedent) and using `Worksheet.iter_rows(min_row=,
max_row=, values_only=True)` to avoid materializing rows outside the
requested window. `data_only=False` is used deliberately (opposite of
Slice 2's discovery, which never reads cell values at all) so a formula
cell's STORED FORMULA TEXT is what gets displayed -- never a cached or
recalculated value, and never a live spreadsheet recalculation (task's
own explicit "do not attempt spreadsheet recalculation"). Excel's own
row/column totals are never re-scanned here at all -- they already exist
on the selected `WorksheetInfo` from Slice 2's own upload-time discovery
(best-effort, exactly as already documented there).

Slice 4 overlay application (`_apply_working_overlay`): a post-
processing step run on the already-fetched raw page, AFTER either
format's own raw-reading logic above -- it never touches how rows are
read, only what gets returned. For each row in the page: any cell
override at `(worksheet_index, row.row_number, column_index)` replaces
that cell's displayed value (the row's own `cells` list is extended
with `None` -- "raw blank" -- only as far as needed to show an override
targeting a column beyond that specific row's own raw width, e.g. a
short/ragged CSV row; every OTHER row's own length is left exactly as
the raw reader produced it). The row's own `excluded` flag is looked up
the same way. Overrides are pre-filtered by worksheet ONCE per preview
call (`_overrides_for_worksheet`), not re-scanned per row, so this stays
proportional to (page size + total edit count for this worksheet),
never to the raw dataset's size.

Slice 5 structure-mapping application (`_apply_structure_mapping`): a
SECOND post-processing step, run after `_apply_working_overlay`, adding
`is_header`/`in_active_region` flags to each row of the page and
computing the page-independent `column_labels`/`column_roles` lists.
Column labels need the header row's own WORKING cells regardless of
which page was actually requested (task's own "header may be on a page
not currently visible" acknowledgment) -- `_resolve_header_cells()`
reuses the already-fetched page when the header row happens to be on
it, and otherwise performs exactly one extra single-row fetch (never a
second full-page read) and applies the SAME overlay-merge logic to that
one row alone, so a working header edit is reflected in
`column_labels` exactly like it is in the row's own `cells`.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from dataclasses import dataclass, field
from typing import Any, Iterator

from openpyxl import load_workbook

from app.domain.channel_classification import ENGINEERING_QUANTITY_UNDEFINED
from app.domain.preparation_session import FORMAT_CSV, FORMAT_EXCEL, PreparationSession
from app.domain.working_overlay import (
    END_MODE_SPECIFIC,
    OVERRIDE_KIND_CLEAR,
    ROLE_NOT_ASSIGNED,
    WorkingOverlay,
)
from app.services.errors import SourceNotFoundError, WorkbookParseError, WorksheetNotSelectedError
from app.services.preparation_session_registry import PreparationSessionRegistry

#: Task's own suggested default; a bounded maximum enforced server-side
#: regardless of what a caller requests (see this module's own docstring
#: note on why offset/limit validation itself lives at the API's
#: `Query(...)` layer, not here).
PREVIEW_DEFAULT_LIMIT = 200
PREVIEW_MAX_LIMIT = 1000

#: How a reported row/column total was obtained -- task's own explicit
#: "clearly document whether they are: exact / best effort / unknown"
#: requirement. CSV counts are ALWAYS "exact" here (a real scan of the
#: actual in-memory bytes, not a sample or estimate); Excel counts
#: inherit Slice 2's own "best_effort" (from `max_row`/`max_column`) or
#: "unknown" (when the workbook's own XML didn't expose a dimension hint
#: at all, see `WorksheetInfo`'s own docstring).
ROW_BASIS_EXACT = "exact"
ROW_BASIS_BEST_EFFORT = "best_effort"
ROW_BASIS_UNKNOWN = "unknown"

#: A conservative, small allowlist for CSV delimiter sniffing -- task's
#: own "do not treat delimiter selection as an engineering interpretation"
#: guardrail. Restricting `csv.Sniffer` to only these candidates (rather
#: than letting it guess freely) is a real, verified fix for a genuine
#: failure mode: an unrestricted sniff on single-column data can return
#: a nonsense "delimiter" (a letter from the data itself) instead of
#: failing cleanly -- verified directly against this exact case before
#: choosing this approach.
_CSV_CANDIDATE_DELIMITERS = ",;\t|"
_CSV_DEFAULT_DELIMITER = ","
_CSV_SNIFF_SAMPLE_CHARS = 8192


@dataclass(slots=True)
class ModifiedCell:
    """One cell within a `PreviewRow` that currently has an active
    working override -- sparse by construction (task's own "do not
    multiply payload size unnecessarily for every unchanged cell"):
    only cells that actually differ from the raw source appear here.
    `raw_value` is the ORIGINAL value at this position, in its native
    type, preserved for provenance/hover/reset display -- never the
    working value (that already lives in the row's own `cells`)."""

    column_index: int
    raw_value: Any


@dataclass(slots=True)
class PreviewRow:
    """One row, as WORKING values (raw with the overlay applied).
    `row_number` is 1-based and matches the source's own row position
    (CSV: `csv.reader`'s own enumeration; Excel: the worksheet's own row
    index) -- never renumbered/reindexed, including when `excluded` is
    `True` or `in_active_region` is `False` (task's own explicit
    "provenance" requirement: both are flags, never a removal or a
    renumbering). `cells` is the DISPLAYED (working) value at each
    position: CSV cells are `str` unless overridden; Excel cells keep
    their native JSON-safe type unless overridden, with `datetime`/
    `date`/`time` raw values converted to ISO-8601 strings purely for
    JSON transport. `modified_cells` lists only the cells in THIS row
    with an active override, each carrying the raw value alongside for
    provenance -- see `ModifiedCell`'s own docstring.

    Slice 5: `is_header` flags the currently-selected header row (at
    most one row per page can ever have this `True`, since a source has
    at most one header row per worksheet). `in_active_region` reflects
    the current data-region narrowing -- `True` for every row when no
    region has been set (Slice 4's own original "entire source is
    active" default, unchanged by Slice 5 unless the user narrows it).
    `excluded` (Slice 4) and `in_active_region` (Slice 5) are
    independent: a row can be inside the active region and still
    excluded, or outside the region and not excluded -- never
    conflated (task's own explicit "these are different concepts"
    guardrail)."""

    row_number: int
    cells: list[Any]
    excluded: bool = False
    modified_cells: list[ModifiedCell] = field(default_factory=list)
    is_header: bool = False
    in_active_region: bool = True


@dataclass(slots=True)
class PreviewResult:
    """See this module's own docstring for the CSV/Excel strategies that
    produce this. `selected_worksheet_index` is `None` for CSV (no
    worksheet concept at all -- never fabricated). `working_revision` is
    `WorkingOverlay.revision` at the moment this page was read, for the
    frontend's own stale-page/refresh bookkeeping (task's own
    "lightweight revision counter... stale-page detection").

    UAT fix (2026-09-04): the Slice 4-era `ignored_columns` field
    (column indices with the now-retired `ignore` role) is retired --
    `column_roles` (Slice 5) already carries the same information for
    EVERY column, not just the ones a caller would once have called
    "ignored," and the three-role simplification (`not_assigned`/
    `time_axis`/`waveform`) collapses "ignored" and "unclassified" into
    the SAME single state anyway, so a separate field would be purely
    redundant. A caller wanting "which columns are not analysed by
    Powerwave" now derives it directly from `column_roles`.

    Slice 5: `header_row_number`/`data_start_row`/`data_end_mode`/
    `data_end_row` mirror
    `app.services.working_overlay_service.WorkingOverlaySummary`'s own
    same-named fields for the worksheet/source this page belongs to
    (`None` when unset). `column_labels`/`column_roles` are each sized
    to `column_count` (empty when `column_count` is unknown) -- small,
    O(columns) payloads, never duplicated per row.

    `data_end_mode`/`data_end_row` (a later owner-UAT refinement) mirror
    `app.domain.working_overlay.DataRegion`'s own two fields verbatim --
    `data_end_row` is `None` for `END_MODE_SOURCE_END` (never a
    resolved/guessed numeric value; see that constant's own docstring).
    The ACTUAL resolved upper bound used to compute each row's own
    `in_active_region` flag is an internal-only detail of
    `_apply_structure_mapping()`, never itself exposed as a separate
    field -- the frontend renders "end" literally for
    `END_MODE_SOURCE_END`, it never needs the resolved number."""

    source_id: str
    selected_worksheet_index: int | None
    offset: int
    limit: int
    returned_row_count: int
    total_row_count: int | None
    total_row_count_basis: str
    column_count: int | None
    column_count_basis: str
    rows: list[PreviewRow]
    working_revision: int = 0
    header_row_number: int | None = None
    data_start_row: int | None = None
    data_end_mode: str | None = None
    data_end_row: int | None = None
    column_labels: list[str] = field(default_factory=list)
    column_roles: list[str] = field(default_factory=list)
    # Engineering Quantity enhancement (DEC-077): one Engineering Quantity
    # per column, sized/aligned exactly like column_labels/column_roles
    # above, defaulting to ENGINEERING_QUANTITY_UNDEFINED for any column
    # with no explicit column_engineering_quantities entry -- the SAME
    # "absence is the default" convention column_roles already uses. The
    # frontend only ever SHOWS this for a column whose role is currently
    # `waveform`; the value is still reported for every column so a prior
    # selection survives a column's role moving away from Waveform and
    # back (see app.domain.working_overlay.set_column_role's own
    # docstring for why nothing here clears it automatically).
    column_engineering_quantities: list[str] = field(default_factory=list)


def _sniff_csv_delimiter(sample: str) -> str:
    """Bounded, conservative dialect sniffing -- never a full-file scan
    just to pick a delimiter, and never treated as an "engineering
    interpretation" (task's own wording): a wrong guess only affects how
    this ONE preview splits cells, never anything downstream, and later
    slices (header/column mapping) start from a clean slate regardless.
    """
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=_CSV_CANDIDATE_DELIMITERS)
        return dialect.delimiter
    except csv.Error:
        return _CSV_DEFAULT_DELIMITER


def _open_csv_reader(session: PreparationSession) -> csv.reader:
    # errors="replace" (never raises): task scope is raw structural
    # preview, not encoding detection -- an undecodable byte becomes a
    # visible replacement character rather than failing the whole
    # preview. This is a disclosed simplification, not a claim that
    # encoding detection is solved.
    text = session.raw_bytes.decode("utf-8", errors="replace")
    delimiter = _sniff_csv_delimiter(text[:_CSV_SNIFF_SAMPLE_CHARS])
    return csv.reader(io.StringIO(text), delimiter=delimiter)


def ensure_csv_totals_cached(session: PreparationSession) -> None:
    """Populate `session.cached_row_count`/`cached_column_count` via one
    full pass over the in-memory text, if not already cached. Shared by
    `_preview_csv()` (below) and `app.services.working_overlay_service`'s
    own coordinate-bounds validation, so there is exactly one CSV
    row/column-counting implementation, never two that could disagree.
    A no-op (zero cost) once already cached."""
    if session.cached_row_count is not None:
        return
    reader = _open_csv_reader(session)
    row_count = 0
    max_columns = 0
    for row_number, row in enumerate(reader, start=1):
        row_count = row_number
        if len(row) > max_columns:
            max_columns = len(row)
    session.cached_row_count = row_count
    session.cached_column_count = max_columns


def _fetch_single_csv_row(session: PreparationSession, row_number: int) -> list[str] | None:
    """Read exactly one raw CSV row (1-based) without materializing the
    rest of the file into the returned page -- used only for a header
    row that falls outside the currently requested page (see this
    module's own docstring). Still an O(row_number) scan (no index
    exists), but bounded to a single row's worth of work beyond that,
    and only ever invoked once per preview request at most."""
    reader = _open_csv_reader(session)
    for i, row in enumerate(reader, start=1):
        if i == row_number:
            return list(row)
        if i > row_number:
            break
    return None


def _fetch_single_excel_row(session: PreparationSession, worksheet_index: int | None, row_number: int) -> list[Any] | None:
    """Excel counterpart of `_fetch_single_csv_row` -- reopens the
    workbook fresh (same "never held open across requests" policy as
    `_preview_excel`) purely to read one row's values."""
    if worksheet_index is None:
        return None
    worksheet_info = session.summary.worksheets[worksheet_index]
    try:
        workbook = load_workbook(io.BytesIO(session.raw_bytes), read_only=True, data_only=False)
    except Exception:
        return None
    try:
        worksheet = workbook[worksheet_info.name]
        for row_values in worksheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True):
            return [_json_safe_excel_cell(v) for v in row_values]
        return None
    finally:
        workbook.close()


def _preview_csv(session: PreparationSession, *, offset: int, limit: int) -> PreviewResult:
    reader = _open_csv_reader(session)

    cached_total = session.cached_row_count
    stop_after = offset + limit  # 1-based inclusive row_number

    page: list[PreviewRow] = []
    row_count = 0
    max_columns = 0

    for row_number, row in enumerate(reader, start=1):
        row_count = row_number
        if len(row) > max_columns:
            max_columns = len(row)
        if offset < row_number <= stop_after:
            page.append(PreviewRow(row_number=row_number, cells=list(row)))
        if cached_total is not None and row_number >= stop_after:
            # Totals are already known from a prior request -- no need
            # to keep scanning once this page is fully collected.
            row_count = cached_total
            max_columns = session.cached_column_count or 0
            break
    else:
        # Loop reached the true end of the file (either cached_total was
        # None, so a full scan was required anyway, or the requested
        # page extends past the last row) -- row_count/max_columns are
        # now the real totals; memoize them for every future request
        # against this same session, regardless of which page they ask
        # for next.
        session.cached_row_count = row_count
        session.cached_column_count = max_columns

    total_row_count = session.cached_row_count if session.cached_row_count is not None else row_count
    column_count = session.cached_column_count if session.cached_column_count is not None else max_columns

    _apply_working_overlay(session, worksheet_index=None, rows=page)
    (
        header_row_number, data_start_row, data_end_mode, data_end_row,
        column_labels, column_roles, column_engineering_quantities,
    ) = _apply_structure_mapping(
        session, worksheet_index=None, page=page, column_count=column_count, known_row_total=total_row_count,
    )

    return PreviewResult(
        source_id=session.summary.source_id,
        selected_worksheet_index=None,
        offset=offset,
        limit=limit,
        returned_row_count=len(page),
        total_row_count=total_row_count,
        total_row_count_basis=ROW_BASIS_EXACT,
        column_count=column_count,
        column_count_basis=ROW_BASIS_EXACT,
        rows=page,
        working_revision=session.working_overlay.revision,
        header_row_number=header_row_number,
        data_start_row=data_start_row,
        data_end_mode=data_end_mode,
        data_end_row=data_end_row,
        column_labels=column_labels,
        column_roles=column_roles,
        column_engineering_quantities=column_engineering_quantities,
    )


def _json_safe_excel_cell(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return value


def _preview_excel(session: PreparationSession, *, offset: int, limit: int) -> PreviewResult:
    summary = session.summary
    if summary.selected_worksheet_index is None:
        raise WorksheetNotSelectedError(
            "This workbook has more than one worksheet; select one with "
            "PATCH .../preparation-sources/{source_id} before requesting a preview."
        )
    worksheet_index = summary.selected_worksheet_index
    worksheet_info = summary.worksheets[worksheet_index]

    try:
        workbook = load_workbook(io.BytesIO(session.raw_bytes), read_only=True, data_only=False)
    except Exception as exc:
        raise WorkbookParseError(f"Could not re-open the Excel workbook for preview: {exc}") from exc

    try:
        worksheet = workbook[worksheet_info.name]
        min_row = offset + 1
        max_row = offset + limit
        page: list[PreviewRow] = []
        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row
        ):
            page.append(
                PreviewRow(
                    row_number=row_number,
                    cells=[_json_safe_excel_cell(v) for v in row_values],
                )
            )
    finally:
        # Never held open across requests -- released before this
        # function returns, whether it succeeded or raised.
        workbook.close()

    _apply_working_overlay(session, worksheet_index=worksheet_index, rows=page)
    column_count = worksheet_info.column_count or 0
    (
        header_row_number, data_start_row, data_end_mode, data_end_row,
        column_labels, column_roles, column_engineering_quantities,
    ) = _apply_structure_mapping(
        session, worksheet_index=worksheet_index, page=page, column_count=column_count,
        known_row_total=worksheet_info.row_count,
    )

    return PreviewResult(
        source_id=summary.source_id,
        selected_worksheet_index=worksheet_index,
        offset=offset,
        limit=limit,
        returned_row_count=len(page),
        total_row_count=worksheet_info.row_count,
        total_row_count_basis=ROW_BASIS_BEST_EFFORT if worksheet_info.row_count is not None else ROW_BASIS_UNKNOWN,
        column_count=worksheet_info.column_count,
        column_count_basis=ROW_BASIS_BEST_EFFORT if worksheet_info.column_count is not None else ROW_BASIS_UNKNOWN,
        rows=page,
        working_revision=session.working_overlay.revision,
        header_row_number=header_row_number,
        data_start_row=data_start_row,
        data_end_mode=data_end_mode,
        data_end_row=data_end_row,
        column_labels=column_labels,
        column_roles=column_roles,
        column_engineering_quantities=column_engineering_quantities,
    )


def _overrides_for_worksheet(overlay: WorkingOverlay, worksheet_index: int | None) -> dict[int, dict[int, Any]]:
    """Pre-filter the WHOLE overlay's cell overrides down to just this
    worksheet, grouped by row -- done ONCE per preview call (not once
    per row), so applying overrides to a page stays O(page size +
    this-worksheet's own edit count), never O(page size x total edit
    count)."""
    by_row: dict[int, dict[int, Any]] = {}
    for (ws, row_number, column_index), override in overlay.cell_overrides.items():
        if ws != worksheet_index:
            continue
        by_row.setdefault(row_number, {})[column_index] = override
    return by_row


def _excluded_rows_for_worksheet(overlay: WorkingOverlay, worksheet_index: int | None) -> set[int]:
    return {row_number for (ws, row_number) in overlay.excluded_rows if ws == worksheet_index}




def _apply_working_overlay(session: PreparationSession, *, worksheet_index: int | None, rows: list[PreviewRow]) -> None:
    """Mutate `rows` (already-fetched RAW rows) in place into their
    WORKING form -- see this module's own docstring for the exact
    strategy. Never touches `session.raw_bytes` or the overlay itself
    (read-only with respect to `session.working_overlay`)."""
    overlay = session.working_overlay
    if not overlay.cell_overrides and not overlay.excluded_rows:
        return  # common case (no edits yet at all) -- skip the filtering work entirely

    overrides_by_row = _overrides_for_worksheet(overlay, worksheet_index)
    excluded_row_numbers = _excluded_rows_for_worksheet(overlay, worksheet_index)

    for row in rows:
        row.excluded = row.row_number in excluded_row_numbers
        row_overrides = overrides_by_row.get(row.row_number)
        if not row_overrides:
            continue
        # Extend this ONE row's own cells with raw-blank (None) padding
        # only as far as needed to show an override targeting a column
        # beyond its own raw width (a short/ragged row) -- every other
        # row's own length is left exactly as the raw reader produced it
        # (task's own "editing a blank raw cell must be supported"
        # requirement, without silently padding every unmodified row).
        needed_len = max(len(row.cells), max(row_overrides) + 1)
        if needed_len > len(row.cells):
            row.cells = row.cells + [None] * (needed_len - len(row.cells))
        modified: list[ModifiedCell] = []
        for column_index in sorted(row_overrides):
            override = row_overrides[column_index]
            raw_value = row.cells[column_index]
            row.cells[column_index] = None if override.kind == OVERRIDE_KIND_CLEAR else override.value
            modified.append(ModifiedCell(column_index=column_index, raw_value=raw_value))
        row.modified_cells = modified


def _spreadsheet_column_label(index: int) -> str:
    """Same A, B, ..., Z, AA, AB, ... scheme the frontend's own
    `wwSpreadsheetColumnLabel()` uses -- kept in sync deliberately (task
    section: "do not destroy the existing A/B/C column coordinate
    references")."""
    label = ""
    n = index
    while n >= 0:
        label = chr(65 + (n % 26)) + label
        n = n // 26 - 1
    return label


def _build_column_labels(header_cells: list[Any] | None, column_count: int) -> list[str]:
    """One display label per column (Slice 5). `header_cells is None`
    means no header row is selected at all -- every label is the plain
    spreadsheet letter (task section: "When no header selected:
    column_labels = spreadsheet-style fallback"). When a header IS
    selected, a genuinely blank cell (raw `None`, or a working edit to
    `""`) gets the distinct `"Column {letter}"` fallback instead (task
    section: "A blank header should receive a safe fallback display
    label") -- so the two fallback cases stay visually distinguishable.
    Duplicate labels are allowed verbatim, deliberately NOT
    disambiguated here (task's own "keep implementation simple and
    stable-index-based" guidance) -- the frontend already shows each
    column's own stable letter alongside its label, which is enough to
    tell duplicates apart without inventing suffixes."""
    labels: list[str] = []
    for c in range(column_count):
        letter = _spreadsheet_column_label(c)
        if header_cells is None:
            labels.append(letter)
            continue
        raw = header_cells[c] if c < len(header_cells) else None
        text = "" if raw is None else str(raw)
        labels.append(text if text != "" else f"Column {letter}")
    return labels


def resolve_single_column_label(
    session: PreparationSession, *, worksheet_index: int | None, column_index: int,
) -> str:
    """One column's current WORKING display label -- the exact same value
    `_build_column_labels()` would produce for this column, without
    requiring a full preview page fetch. Used by
    `app.services.working_overlay_service`'s own Engineering-Quantity-
    suffix-restoration step (DEC-077, `set_column_role()`) -- reuses the
    SAME header-resolution/fallback logic `_apply_structure_mapping()`
    already established (`_resolve_header_cells()`/`_build_column_labels()`),
    never a second, independent label-lookup implementation. Passing an
    empty `page=[]` forces `_resolve_header_cells()` down its own
    single-row-fetch fallback path (the header row is virtually never
    already in hand here, since this is called from a column-role
    mutation, not a preview read)."""
    overlay = session.working_overlay
    header_row_number = overlay.header_row.get(worksheet_index)
    header_cells = _resolve_header_cells(
        session, worksheet_index=worksheet_index, header_row_number=header_row_number, page=[],
    )
    return _build_column_labels(header_cells, column_index + 1)[column_index]


def _build_column_roles(overlay: WorkingOverlay, worksheet_index: int | None, column_count: int) -> list[str]:
    """One role per column (Slice 5), defaulting to `ROLE_NOT_ASSIGNED`
    for any column with no explicit `column_roles` entry -- the model's
    own "absence is the default, never automatically classified"
    guarantee (see `app.domain.working_overlay`'s own module
    docstring), made visible here as an explicit value per column
    rather than a sparse dict the frontend would have to fill in
    itself."""
    return [
        overlay.column_roles.get((worksheet_index, c), ROLE_NOT_ASSIGNED)
        for c in range(column_count)
    ]


def _build_column_engineering_quantities(
    overlay: WorkingOverlay, worksheet_index: int | None, column_count: int,
) -> list[str]:
    """One Engineering Quantity per column (DEC-077), defaulting to
    `ENGINEERING_QUANTITY_UNDEFINED` for any column with no explicit
    `column_engineering_quantities` entry -- mirrors `_build_column_roles()`
    above exactly, same sparse-dict-to-dense-list shape."""
    return [
        overlay.column_engineering_quantities.get((worksheet_index, c), ENGINEERING_QUANTITY_UNDEFINED)
        for c in range(column_count)
    ]


def _resolve_header_cells(
    session: PreparationSession,
    *,
    worksheet_index: int | None,
    header_row_number: int | None,
    page: list[PreviewRow],
) -> list[Any] | None:
    """Return the header row's WORKING cells, or `None` if no header is
    selected. Reuses the row from `page` (already overlay-applied) when
    the header happens to fall within the current page's own window --
    the common case for a header near the top of a file being previewed
    from its own first page; otherwise performs exactly one extra
    single-row fetch (`_fetch_single_csv_row`/`_fetch_single_excel_row`)
    and applies the SAME overlay-merge logic
    (`_apply_working_overlay`) to that one row alone, so a working
    header edit is reflected here exactly like it is in that row's own
    `cells` when it IS on the current page (task section: "Header row
    and working edits")."""
    if header_row_number is None:
        return None
    for row in page:
        if row.row_number == header_row_number:
            return row.cells
    if session.summary.source_format == FORMAT_CSV:
        raw_cells = _fetch_single_csv_row(session, header_row_number)
    else:
        raw_cells = _fetch_single_excel_row(session, worksheet_index, header_row_number)
    if raw_cells is None:
        return None
    header_row = PreviewRow(row_number=header_row_number, cells=list(raw_cells))
    _apply_working_overlay(session, worksheet_index=worksheet_index, rows=[header_row])
    return header_row.cells


def _apply_structure_mapping(
    session: PreparationSession,
    *,
    worksheet_index: int | None,
    page: list[PreviewRow],
    column_count: int,
    known_row_total: int | None,
) -> tuple[int | None, int | None, str | None, int | None, list[str], list[str], list[str]]:
    """Slice 5's own post-processing step, run after
    `_apply_working_overlay` -- adds `is_header`/`in_active_region` to
    every row of `page` (mutated in place, same convention as that
    function) and returns the page-independent
    `(header_row_number, data_start_row, data_end_mode, data_end_row,
    column_labels, column_roles)` tuple for the caller's own
    `PreviewResult`. Never touches `cells`/`modified_cells`/`excluded`
    -- those stay exactly as `_apply_working_overlay` left them.

    `known_row_total` (a later owner-UAT refinement) is this source/
    worksheet's own already-known row total -- exact for CSV, best-
    effort or `None` for Excel -- passed in by the caller rather than
    re-derived here, since both callers already have it on hand from
    their own raw-reading pass. It is used ONLY to resolve
    `END_MODE_SOURCE_END`'s own floating upper bound for the
    `in_active_region` flag; the reported `data_end_row` for that mode
    stays `None` regardless (never a resolved/guessed value -- see
    `app.domain.working_overlay.DataRegion`'s own docstring)."""
    overlay = session.working_overlay
    header_row_number = overlay.header_row.get(worksheet_index)
    region = overlay.data_region.get(worksheet_index)
    data_start_row = region.start_row if region else None
    data_end_mode = region.end_mode if region else None
    data_end_row = region.end_row if region else None

    if region is None:
        resolved_end = None  # no region configured at all -- entire source active
    elif region.end_mode == END_MODE_SPECIFIC:
        resolved_end = region.end_row
    else:  # END_MODE_SOURCE_END -- floats with whatever total is actually known
        resolved_end = known_row_total

    for row in page:
        row.is_header = header_row_number is not None and row.row_number == header_row_number
        row.in_active_region = (
            (data_start_row is None or row.row_number >= data_start_row)
            and (resolved_end is None or row.row_number <= resolved_end)
        )

    header_cells = _resolve_header_cells(
        session, worksheet_index=worksheet_index, header_row_number=header_row_number, page=page,
    )
    column_labels = _build_column_labels(header_cells, column_count) if column_count else []
    column_roles = _build_column_roles(overlay, worksheet_index, column_count) if column_count else []
    column_engineering_quantities = (
        _build_column_engineering_quantities(overlay, worksheet_index, column_count) if column_count else []
    )
    return (
        header_row_number, data_start_row, data_end_mode, data_end_row,
        column_labels, column_roles, column_engineering_quantities,
    )


def preview_preparation_source(
    *,
    workspace_id: str,
    source_id: str,
    offset: int,
    limit: int,
    registry: PreparationSessionRegistry,
) -> PreviewResult:
    """Return one bounded page of WORKING rows (raw + overlay applied)
    for a CSV or Excel preparation source. `offset`/`limit` are
    trusted here to already be within bounds (see this module's own
    docstring: enforced by the API's own `Query(ge=..., le=...)`
    constraints, matching this codebase's existing `point_budget`
    precedent) -- this function does not re-validate them a second time.

    Raises `SourceNotFoundError` if no such preparation session exists,
    `WorksheetNotSelectedError` for an Excel source with no worksheet
    chosen yet, or `WorkbookParseError` if the stored workbook bytes
    fail to re-open (should not happen in practice -- defensive only).
    """
    session = registry.get(workspace_id, source_id)
    if session is None:
        raise SourceNotFoundError(
            f"No preparation source '{source_id}' in workspace '{workspace_id}'."
        )

    if session.summary.source_format == FORMAT_CSV:
        return _preview_csv(session, offset=offset, limit=limit)
    # FORMAT_EXCEL is the only other known format (see
    # app.domain.preparation_session.KNOWN_PREPARATION_FORMATS) --
    # nothing else can have reached this registry.
    return _preview_excel(session, offset=offset, limit=limit)


def iterate_active_region_rows(
    session: PreparationSession, *, worksheet_index: int | None,
) -> Iterator[PreviewRow]:
    """(Slice 9, DEC-072) Single-PASS streaming iterator over this
    worksheet/source's own CURRENT active data region, in original row
    order, with the SAME working-overlay application (cell overrides,
    row exclusion) and `is_header`/`in_active_region` flags
    `preview_preparation_source()` itself computes -- but never
    materializing more than one row at a time, and never re-scanning
    the source once per page the way repeated `preview_preparation_
    source()` calls would. This is `app.services.readiness_service`'s
    own row source for the full-active-region validation checks Slice 9
    requires (a bounded ≤1000-row PAGE is this module's OWN job, never
    a substitute for a genuine readiness gate -- see that service
    module's own docstring for exactly why).

    Rows strictly before the region's own start are never yielded at
    all (task section P: "do not scan or validate unused footer/header
    rows merely because they exist in the raw source") -- CSV still has
    to read past them sequentially (no index exists; the same accepted
    tradeoff `_fetch_single_csv_row()` already documents) but does no
    override/flag work for them. The loop stops as soon as it passes a
    KNOWN, explicit `END_MODE_SPECIFIC` end row; `END_MODE_SOURCE_END`
    (or no region at all) reads through to the source's own true end,
    since that IS the resolved boundary in that case. The header row
    (if configured) and excluded rows ARE yielded (their own flags set)
    -- exactly like `_fetch_time_axis_samples()`'s own established
    convention, this one iterator lets each caller filter for its own
    slightly different need rather than hard-coding one skip policy
    here.
    """
    overlay = session.working_overlay
    region = overlay.data_region.get(worksheet_index)
    start_row = region.start_row if region else 1
    specific_end = region.end_row if (region and region.end_mode == END_MODE_SPECIFIC) else None
    header_row_number = overlay.header_row.get(worksheet_index)

    overrides_by_row = _overrides_for_worksheet(overlay, worksheet_index)
    excluded_row_numbers = _excluded_rows_for_worksheet(overlay, worksheet_index)

    def _finalize(row: PreviewRow) -> PreviewRow:
        row.excluded = row.row_number in excluded_row_numbers
        row.is_header = header_row_number is not None and row.row_number == header_row_number
        row.in_active_region = row.row_number >= start_row and (specific_end is None or row.row_number <= specific_end)
        row_overrides = overrides_by_row.get(row.row_number)
        if row_overrides:
            needed_len = max(len(row.cells), max(row_overrides) + 1)
            if needed_len > len(row.cells):
                row.cells = row.cells + [None] * (needed_len - len(row.cells))
            modified: list[ModifiedCell] = []
            for column_index in sorted(row_overrides):
                override = row_overrides[column_index]
                raw_value = row.cells[column_index]
                row.cells[column_index] = None if override.kind == OVERRIDE_KIND_CLEAR else override.value
                modified.append(ModifiedCell(column_index=column_index, raw_value=raw_value))
            row.modified_cells = modified
        return row

    if session.summary.source_format == FORMAT_CSV:
        reader = _open_csv_reader(session)
        for row_number, raw_row in enumerate(reader, start=1):
            if row_number < start_row:
                continue
            if specific_end is not None and row_number > specific_end:
                break
            yield _finalize(PreviewRow(row_number=row_number, cells=list(raw_row)))
        return

    # FORMAT_EXCEL
    worksheet_info = session.summary.worksheets[worksheet_index]
    try:
        workbook = load_workbook(io.BytesIO(session.raw_bytes), read_only=True, data_only=False)
    except Exception as exc:
        raise WorkbookParseError(f"Could not re-open the Excel workbook for readiness validation: {exc}") from exc
    try:
        worksheet = workbook[worksheet_info.name]
        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=start_row, max_row=specific_end, values_only=True), start=start_row,
        ):
            yield _finalize(PreviewRow(row_number=row_number, cells=[_json_safe_excel_cell(v) for v in row_values]))
    finally:
        workbook.close()
