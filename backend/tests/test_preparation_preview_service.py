"""Service-level tests for paged raw-data preview (Slice 3, DEC-072).

`preview_preparation_source` is a plain synchronous function (unlike the
upload services, which are async to match FastAPI's `UploadFile`
handling) -- no `asyncio.run()` wrapper needed here.
"""

from __future__ import annotations

import asyncio
import io

import pytest
from fastapi import UploadFile
from openpyxl import Workbook
from starlette.datastructures import Headers

from app.services.errors import SourceNotFoundError, WorksheetNotSelectedError
from app.services.preparation_import_service import (
    import_csv_preparation_source,
    import_excel_preparation_source,
    select_preparation_worksheet,
)
from app.services.preparation_preview_service import (
    ROW_BASIS_BEST_EFFORT,
    ROW_BASIS_EXACT,
    preview_preparation_source,
)
from app.services.preparation_session_registry import PreparationSessionRegistry
from app.services.working_overlay_service import (
    clear_header_row,
    edit_cell,
    reset_all_working_changes,
    reset_cell,
    reset_data_region,
    set_column_role,
    set_data_region,
    set_header_row,
    set_row_excluded,
)


def _upload(content: bytes, filename: str, content_type: str) -> UploadFile:
    return UploadFile(file=io.BytesIO(content), filename=filename, headers=Headers({"content-type": content_type}))


def _add_csv(registry: PreparationSessionRegistry, content: bytes, workspace_id: str = "ws-1", filename: str = "e.csv") -> str:
    summary = asyncio.run(
        import_csv_preparation_source(
            workspace_id=workspace_id, csv_upload=_upload(content, filename, "text/csv"),
            max_total_bytes=100 * 1024 * 1024, registry=registry,
        )
    )
    return summary.source_id


def _build_xlsx(sheets: dict | None = None, hidden: frozenset = frozenset(), formulas: dict | None = None) -> bytes:
    if sheets is None:
        sheets = {"Sheet1": [["a", "b"], [1, 2]]}
    workbook = Workbook()
    names = list(sheets.keys())
    workbook.active.title = names[0]
    for row in sheets[names[0]]:
        workbook.active.append(row)
    for name in names[1:]:
        ws = workbook.create_sheet(name)
        for row in sheets[name]:
            ws.append(row)
    for name in hidden:
        workbook[name].sheet_state = "hidden"
    if formulas:
        for cell_ref, formula in formulas.items():
            workbook.active[cell_ref] = formula
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _add_excel(registry: PreparationSessionRegistry, content: bytes, workspace_id: str = "ws-1", filename: str = "e.xlsx") -> str:
    summary = asyncio.run(
        import_excel_preparation_source(
            workspace_id=workspace_id,
            excel_upload=_upload(content, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            max_total_bytes=100 * 1024 * 1024, registry=registry,
        )
    )
    return summary.source_id


# ---- CSV preview ----


class TestCsvPreview:
    def test_first_page(self):
        registry = PreparationSessionRegistry()
        content = "\n".join(f"{i},v{i}" for i in range(1, 11)).encode()
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=5, registry=registry)

        assert result.offset == 0
        assert result.limit == 5
        assert result.returned_row_count == 5
        assert [r.row_number for r in result.rows] == [1, 2, 3, 4, 5]
        assert result.rows[0].cells == ["1", "v1"]
        assert result.total_row_count == 10
        assert result.total_row_count_basis == ROW_BASIS_EXACT
        assert result.column_count == 2
        assert result.column_count_basis == ROW_BASIS_EXACT
        assert result.selected_worksheet_index is None

    def test_later_page(self):
        registry = PreparationSessionRegistry()
        content = "\n".join(f"{i},v{i}" for i in range(1, 11)).encode()
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=5, limit=5, registry=registry)

        assert [r.row_number for r in result.rows] == [6, 7, 8, 9, 10]
        assert result.rows[0].cells == ["6", "v6"]
        assert result.total_row_count == 10

    def test_partial_final_page(self):
        registry = PreparationSessionRegistry()
        content = "\n".join(f"{i},v{i}" for i in range(1, 11)).encode()  # 10 rows
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=8, limit=5, registry=registry)

        # Only rows 9-10 exist beyond offset 8.
        assert [r.row_number for r in result.rows] == [9, 10]
        assert result.returned_row_count == 2
        assert result.total_row_count == 10

    def test_offset_past_end_returns_empty_page_not_an_error(self):
        registry = PreparationSessionRegistry()
        content = b"a,b\n1,2\n"  # 2 rows total
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=100, limit=10, registry=registry)

        assert result.rows == []
        assert result.returned_row_count == 0
        assert result.total_row_count == 2

    def test_blank_cells_preserved(self):
        registry = PreparationSessionRegistry()
        content = b"a,,c\n,,\n"
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["a", "", "c"]
        assert result.rows[1].cells == ["", "", ""]

    def test_blank_rows_preserved_with_row_number(self):
        registry = PreparationSessionRegistry()
        content = b"a,b\n\n1,2\n"
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert [r.row_number for r in result.rows] == [1, 2, 3]
        assert result.rows[1].cells == []  # the blank row itself
        assert result.rows[2].cells == ["1", "2"]

    def test_many_columns(self):
        registry = PreparationSessionRegistry()
        wide_row = ",".join(str(i) for i in range(50))
        content = (wide_row + "\n").encode()
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_count == 50
        assert len(result.rows[0].cells) == 50

    def test_first_row_is_not_treated_as_a_header(self):
        registry = PreparationSessionRegistry()
        content = b"time,VA\n0.0,1.0\n"
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        # Row 1 (the would-be "header") is returned as an ordinary raw
        # row, not consumed/skipped/relabeled.
        assert result.rows[0].row_number == 1
        assert result.rows[0].cells == ["time", "VA"]

    def test_deleted_source_cannot_preview(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n")
        registry.remove("ws-1", source_id)

        with pytest.raises(SourceNotFoundError):
            preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

    def test_unknown_source_raises_source_not_found(self):
        registry = PreparationSessionRegistry()

        with pytest.raises(SourceNotFoundError):
            preview_preparation_source(workspace_id="ws-1", source_id="does-not-exist", offset=0, limit=10, registry=registry)

    def test_totals_are_cached_after_first_preview_and_reused(self):
        registry = PreparationSessionRegistry()
        content = "\n".join(f"{i},v{i}" for i in range(1, 11)).encode()
        source_id = _add_csv(registry, content)

        preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=3, registry=registry)
        session = registry.get("ws-1", source_id)
        assert session.cached_row_count == 10
        assert session.cached_column_count == 2

        # A later page still reports the same correct total without
        # needing to re-derive it from scratch (implementation detail,
        # but observable via the correct total on a page that itself
        # doesn't reach the end of the file).
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=3, limit=2, registry=registry)
        assert result.total_row_count == 10


class TestCsvStructure:
    def test_comma_delimited(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b,c\n1,2,3\n")

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["a", "b", "c"]

    def test_quoted_commas_are_not_split(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b'a,"b,c",d\n1,2,3\n')

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["a", "b,c", "d"]

    def test_semicolon_delimiter_is_sniffed(self):
        registry = PreparationSessionRegistry()
        content = b"a;b;c\n1;2;3\n4;5;6\n"
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["a", "b", "c"]
        assert result.column_count == 3

    def test_ambiguous_single_column_falls_back_to_comma_default(self):
        # No real delimiter present anywhere -- sniffing must fail
        # cleanly and fall back to the safe default, never guess a
        # nonsense delimiter from the data itself.
        registry = PreparationSessionRegistry()
        content = b"onlyonecolumn\nanothervalue\n"
        source_id = _add_csv(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["onlyonecolumn"]
        assert result.rows[1].cells == ["anothervalue"]

    def test_no_trailing_newline_still_reads_the_last_row(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2")  # no trailing \n

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert [r.cells for r in result.rows] == [["a", "b"], ["1", "2"]]


# ---- Excel preview ----


class TestExcelPreview:
    def test_one_sheet_workbook_preview(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"Only": [["time", "VA"], [0.0, 1.0], [0.001, 2.0]]})
        source_id = _add_excel(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.selected_worksheet_index == 0  # auto-selected, Slice 2
        assert [r.row_number for r in result.rows] == [1, 2, 3]
        assert result.rows[0].cells == ["time", "VA"]
        assert result.total_row_count == 3
        assert result.total_row_count_basis == ROW_BASIS_BEST_EFFORT
        assert result.column_count == 2

    def test_multi_sheet_requires_selection_first(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"A": [["x"]], "B": [["y"]]})
        source_id = _add_excel(registry, content)

        with pytest.raises(WorksheetNotSelectedError):
            preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

    def test_multi_sheet_preview_after_explicit_selection(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"Event Data": [["e1"]], "RMS": [["r1"], ["r2"]]})
        source_id = _add_excel(registry, content)
        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=1, registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.selected_worksheet_index == 1
        assert [r.cells for r in result.rows] == [["r1"], ["r2"]]

    def test_switching_worksheet_changes_the_preview(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"A": [["from-a"]], "B": [["from-b"]]})
        source_id = _add_excel(registry, content)

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=0, registry=registry)
        result_a = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert result_a.rows[0].cells == ["from-a"]

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=1, registry=registry)
        result_b = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert result_b.rows[0].cells == ["from-b"]

    def test_hidden_sheet_can_still_be_previewed_once_selected(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"Visible": [["v"]], "Hidden": [["h"]]}, hidden=frozenset({"Hidden"}))
        source_id = _add_excel(registry, content)

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=1, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["h"]

    def test_blank_cells_are_none(self):
        registry = PreparationSessionRegistry()
        workbook = Workbook()
        workbook.active.append(["a", None, "c"])
        buf = io.BytesIO()
        workbook.save(buf)
        source_id = _add_excel(registry, buf.getvalue())

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["a", None, "c"]

    def test_formula_cell_shows_stored_formula_text(self):
        registry = PreparationSessionRegistry()
        workbook = Workbook()
        ws = workbook.active
        ws["A1"] = 5
        ws["A2"] = "=A1+1"
        buf = io.BytesIO()
        workbook.save(buf)
        source_id = _add_excel(registry, buf.getvalue())

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == [5]
        assert result.rows[1].cells == ["=A1+1"]

    def test_partial_final_page(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"Only": [[i] for i in range(1, 6)]})  # 5 rows
        source_id = _add_excel(registry, content)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=3, limit=10, registry=registry)

        assert [r.row_number for r in result.rows] == [4, 5]

    def test_worksheet_resources_are_released_after_preview(self):
        # No direct handle-leak assertion possible from outside openpyxl,
        # but this confirms repeated previews against the same session
        # never fail/hang, consistent with the workbook being reopened
        # and closed cleanly every call (see the service's own docstring).
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"Only": [["x"]]})
        source_id = _add_excel(registry, content)

        for _ in range(5):
            preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)


# ---- CSV/Excel ingestion Slice 4 (DEC-072): working overlay merged into preview ----


class TestWorkingOverlayInCsvPreview:
    def test_edited_cell_shows_the_working_value(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n")

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=2, column_index=0, value="EDITED", registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[1].cells == ["EDITED", "2"]

    def test_edited_cell_reports_provenance_via_modified_cells(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n")

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=2, column_index=0, value="EDITED", registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        edited_row = result.rows[1]
        assert len(edited_row.modified_cells) == 1
        assert edited_row.modified_cells[0].column_index == 0
        assert edited_row.modified_cells[0].raw_value == "1"

    def test_unmodified_rows_have_no_modified_cells(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n")

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=2, column_index=0, value="EDITED", registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].modified_cells == []

    def test_cleared_cell_shows_none(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n")

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=1, value=None, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["a", None]

    def test_reset_cell_restores_the_raw_value_in_preview(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n")

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="X", registry=registry)
        reset_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["a", "b"]
        assert result.rows[0].modified_cells == []

    def test_editing_a_blank_cell_beyond_a_ragged_rows_own_width_is_supported(self):
        # Row 2 only has one column ("x") -- editing column_index=2 on it
        # must pad ONLY this row, not silently widen every other row.
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b,c\nx\n")

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=2, column_index=2, value="NEW", registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[1].cells == ["x", None, "NEW"]
        assert result.rows[0].cells == ["a", "b", "c"]  # untouched row keeps its own original width

    def test_excluded_row_is_flagged_but_never_removed_or_renumbered(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n5,6\n")  # rows 1-4

        set_row_excluded(workspace_id="ws-1", source_id=source_id, row_number=3, excluded=True, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert [r.row_number for r in result.rows] == [1, 2, 3, 4]
        assert [r.excluded for r in result.rows] == [False, False, True, False]
        assert result.rows[2].cells == ["3", "4"]  # excluded row's own data is unchanged

    def test_not_assigned_column_is_reported_page_independently(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b,c\n1,2,3\n")  # row 1: a,b,c / row 2: 1,2,3

        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=1, role="waveform", registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_roles == ["not_assigned", "waveform", "not_assigned"]
        # A column's role never removes it from the returned cells --
        # it is a classification, not a structural deletion.
        assert result.rows[1].cells == ["1", "2", "3"]

    def test_working_revision_reflects_the_overlays_own_revision_counter(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n")

        before = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert before.working_revision == 0

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="X", registry=registry)
        after = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert after.working_revision == 1

    def test_reset_all_restores_the_full_raw_preview(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n")  # 3 rows

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="X", registry=registry)
        set_row_excluded(workspace_id="ws-1", source_id=source_id, row_number=2, excluded=True, registry=registry)
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=1, role="waveform", registry=registry)

        reset_all_working_changes(workspace_id="ws-1", source_id=source_id, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["a", "b"]
        assert [r.excluded for r in result.rows] == [False, False, False]
        assert result.column_roles == ["not_assigned", "not_assigned"]

    def test_no_overlay_activity_skips_overlay_processing_entirely(self):
        # A freshly uploaded/unedited source's preview must look byte-
        # for-byte identical to Slice 3's own raw preview (regression
        # guard for the "common case" early-return in
        # _apply_working_overlay).
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n")

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].excluded is False
        assert result.rows[0].modified_cells == []
        assert result.working_revision == 0

    def test_raw_bytes_are_never_mutated_by_an_edit(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n")
        session = registry.get("ws-1", source_id)
        original_raw_bytes = session.raw_bytes

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="X", registry=registry)

        assert session.raw_bytes == original_raw_bytes
        assert session.raw_bytes is original_raw_bytes


class TestWorkingOverlayInExcelPreview:
    def test_edit_on_selected_worksheet_shows_in_preview(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"A": [["x"]], "B": [["y"]]})
        source_id = _add_excel(registry, content)
        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=0, registry=registry)

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="EDITED", registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.rows[0].cells == ["EDITED"]

    def test_edits_do_not_leak_across_worksheets_in_preview(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"A": [["x"]], "B": [["y"]]})
        source_id = _add_excel(registry, content)

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=0, registry=registry)
        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="from-a", registry=registry)

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=1, registry=registry)
        result_b = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert result_b.rows[0].cells == ["y"]  # untouched -- the edit targeted worksheet 0 only

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=0, registry=registry)
        result_a = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert result_a.rows[0].cells == ["from-a"]



# ---- CSV/Excel ingestion Slice 5 (DEC-072): header/data-region/column-role preview integration ----


class TestHeaderInPreview:
    def test_header_row_flag_and_column_labels(self):
        registry = PreparationSessionRegistry()
        content = b"Station: GPTH\nEvent: Trip\nTime,VR,VY,VB\n0.0,1,2,3\n"
        source_id = _add_csv(registry, content)

        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=3, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.header_row_number == 3
        assert result.column_labels == ["Time", "VR", "VY", "VB"]
        flags = {r.row_number: r.is_header for r in result.rows}
        assert flags == {1: False, 2: False, 3: True, 4: False}
        # Preamble rows remain preserved, not deleted/skipped.
        assert result.rows[0].cells == ["Station: GPTH"]
        assert result.rows[1].cells == ["Event: Trip"]

    def test_clear_header_reverts_to_spreadsheet_letters(self):
        registry = PreparationSessionRegistry()
        content = b"Time,VR\n0.0,1\n"
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)

        clear_header_row(workspace_id="ws-1", source_id=source_id, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.header_row_number is None
        assert result.column_labels == ["A", "B"]
        assert all(not r.is_header for r in result.rows)

    def test_no_header_selected_gives_spreadsheet_letter_labels(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b,c\n1,2,3\n")

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_labels == ["A", "B", "C"]

    def test_blank_header_cell_gets_fallback_label(self):
        registry = PreparationSessionRegistry()
        content = b"Time,VR,,VR\n0.0,1,2,4\n"
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_labels == ["Time", "VR", "Column C", "VR"]

    def test_duplicate_header_labels_are_allowed_verbatim(self):
        registry = PreparationSessionRegistry()
        content = b"Voltage,Voltage,Voltage\n1,2,3\n"
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_labels == ["Voltage", "Voltage", "Voltage"]

    def test_header_row_on_a_different_page_still_resolves_labels(self):
        registry = PreparationSessionRegistry()
        # Build a file where the header (row 1) is NOT on the requested page (offset=10).
        content = ("Time,VR\n" + "\n".join(f"{i},{i}" for i in range(1, 20))).encode()
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=10, limit=5, registry=registry)

        assert result.header_row_number == 1
        assert result.column_labels == ["Time", "VR"]
        assert all(r.row_number != 1 for r in result.rows)  # header itself is not on this page

    def test_working_edit_on_header_cell_updates_the_label(self):
        registry = PreparationSessionRegistry()
        content = b"Vr,Vy\n1,2\n"
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)
        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="VR", registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_labels[0] == "VR"

    def test_resetting_the_header_cell_restores_the_original_label(self):
        registry = PreparationSessionRegistry()
        content = b"Vr,Vy\n1,2\n"
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)
        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="VR", registry=registry)

        reset_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_labels[0] == "Vr"

    def test_header_selected_but_worksheet_starts_unconfigured_for_a_different_sheet(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"A": [["Time", "VR"], [0.0, 1.0]], "B": [["x"], [2]]})
        source_id = _add_excel(registry, content)

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=0, registry=registry)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=1, registry=registry)
        result_b = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert result_b.header_row_number is None
        assert result_b.column_labels == ["A"]

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=0, registry=registry)
        result_a = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert result_a.header_row_number == 1
        assert result_a.column_labels == ["Time", "VR"]


class TestDataRegionInPreview:
    def test_default_region_is_the_entire_source(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n")

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.data_start_row is None
        assert result.data_end_row is None
        assert all(r.in_active_region for r in result.rows)

    def test_rows_outside_region_are_flagged_not_removed(self):
        registry = PreparationSessionRegistry()
        content = b"Station\nEvent\nTime,VR\n0.0,1\n0.001,2\n"
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=3, registry=registry)
        set_data_region(workspace_id="ws-1", source_id=source_id, start_row=4, end_row=5, registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert [r.row_number for r in result.rows] == [1, 2, 3, 4, 5]  # nothing removed
        flags = {r.row_number: r.in_active_region for r in result.rows}
        assert flags == {1: False, 2: False, 3: False, 4: True, 5: True}

    def test_reset_region_reactivates_the_full_source(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n5,6\n")
        set_data_region(workspace_id="ws-1", source_id=source_id, start_row=2, end_row=2, registry=registry)

        reset_data_region(workspace_id="ws-1", source_id=source_id, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.data_start_row is None
        assert all(r.in_active_region for r in result.rows)

    def test_excluded_row_inside_region_is_both_flags_independently(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n5,6\n")
        set_data_region(workspace_id="ws-1", source_id=source_id, start_row=1, end_row=3, registry=registry)
        set_row_excluded(workspace_id="ws-1", source_id=source_id, row_number=2, excluded=True, registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        row2 = next(r for r in result.rows if r.row_number == 2)
        assert row2.in_active_region is True
        assert row2.excluded is True  # independent concepts, both true at once


class TestDataRegionEndModeInPreview:
    """Owner-UAT refinement: `end_mode="source_end"` lets the region's
    own upper bound float with the source's own known total rather than
    requiring a manually-found numeric row. `data_end_row` stays `None`
    on the wire for this mode -- never a resolved/guessed value -- while
    `in_active_region` is still computed correctly using the source's
    own actual (CSV: exact; Excel: best-effort) row total internally."""

    def test_source_end_mode_reports_null_end_row_but_correct_active_flags(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n5,6\n")  # rows 1-4
        set_data_region(workspace_id="ws-1", source_id=source_id, start_row=2, end_mode="source_end", registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.data_start_row == 2
        assert result.data_end_mode == "source_end"
        assert result.data_end_row is None  # never a resolved/guessed numeric value
        flags = {r.row_number: r.in_active_region for r in result.rows}
        assert flags == {1: False, 2: True, 3: True, 4: True}  # correctly resolved to the true last row (4)

    def test_specific_mode_still_reports_the_stored_end_row(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n3,4\n")
        set_data_region(workspace_id="ws-1", source_id=source_id, start_row=1, end_row=2, registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.data_end_mode == "specific"
        assert result.data_end_row == 2

    def test_no_region_at_all_reports_null_end_mode(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n")

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.data_end_mode is None
        assert result.data_end_row is None

    def test_source_end_mode_on_excel_uses_best_effort_worksheet_total(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"Only": [["1"], ["2"], ["3"]]})
        source_id = _add_excel(registry, content)
        set_data_region(workspace_id="ws-1", source_id=source_id, start_row=2, end_mode="source_end", registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.data_end_mode == "source_end"
        assert result.data_end_row is None
        flags = {r.row_number: r.in_active_region for r in result.rows}
        assert flags == {1: False, 2: True, 3: True}

    def test_different_column_lengths_do_not_change_region_semantics(self):
        # A ragged CSV (columns of different effective lengths) must
        # still resolve ONE dataset-wide end -- never a per-column one.
        registry = PreparationSessionRegistry()
        content = b"a,b,c\n1,2,3\n4,5\n6\n"  # row 4 ("6") is much shorter than row 2
        source_id = _add_csv(registry, content)
        set_data_region(workspace_id="ws-1", source_id=source_id, start_row=1, end_mode="source_end", registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        # The resolved end is the source's own total row count (4),
        # regardless of any individual row/column being shorter.
        assert all(r.in_active_region for r in result.rows)
        assert result.total_row_count == 4


class TestColumnRolesInPreview:
    def test_default_roles_are_not_assigned(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b,c\n1,2,3\n")

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_roles == ["not_assigned", "not_assigned", "not_assigned"]

    def test_assigned_roles_appear_in_preview(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b,c,d\n1,2,3,4\n")
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=0, role="time_axis", registry=registry)
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=1, role="waveform", registry=registry)
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=2, role="waveform", registry=registry)
        # column_index=3 is left at its default (not_assigned)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_roles == ["time_axis", "waveform", "waveform", "not_assigned"]

    def test_multiple_time_axis_columns_both_reported(self):
        registry = PreparationSessionRegistry()
        source_id = _add_csv(registry, b"a,b\n1,2\n")
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=0, role="time_axis", registry=registry)
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=1, role="time_axis", registry=registry)

        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_roles == ["time_axis", "time_axis"]

    def test_role_survives_paging(self):
        registry = PreparationSessionRegistry()
        content = ("a,b\n" + "\n".join(f"{i},{i}" for i in range(1, 20))).encode()
        source_id = _add_csv(registry, content)
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=1, role="waveform", registry=registry)

        page1 = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=5, registry=registry)
        page2 = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=10, limit=5, registry=registry)

        assert page1.column_roles == ["not_assigned", "waveform"]
        assert page2.column_roles == ["not_assigned", "waveform"]

    def test_header_edit_does_not_change_role(self):
        registry = PreparationSessionRegistry()
        content = b"Vr,Vy\n1,2\n"
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=0, role="waveform", registry=registry)

        edit_cell(workspace_id="ws-1", source_id=source_id, row_number=1, column_index=0, value="VR", registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.column_labels[0] == "VR"
        assert result.column_roles[0] == "waveform"

    def test_roles_isolated_per_worksheet_in_preview(self):
        registry = PreparationSessionRegistry()
        content = _build_xlsx({"A": [["1"]], "B": [["2"]]})
        source_id = _add_excel(registry, content)

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=0, registry=registry)
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=0, role="waveform", registry=registry)

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=1, registry=registry)
        result_b = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert result_b.column_roles == ["not_assigned"]

        select_preparation_worksheet(workspace_id="ws-1", source_id=source_id, worksheet_index=0, registry=registry)
        result_a = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)
        assert result_a.column_roles == ["waveform"]


class TestResetAllIncludesStructureMapping:
    def test_reset_all_clears_header_region_and_roles_from_preview(self):
        registry = PreparationSessionRegistry()
        content = b"Time,VR\n0.0,1\n0.001,2\n"
        source_id = _add_csv(registry, content)
        set_header_row(workspace_id="ws-1", source_id=source_id, row_number=1, registry=registry)
        set_data_region(workspace_id="ws-1", source_id=source_id, start_row=2, end_row=2, registry=registry)
        set_column_role(workspace_id="ws-1", source_id=source_id, column_index=1, role="waveform", registry=registry)

        reset_all_working_changes(workspace_id="ws-1", source_id=source_id, registry=registry)
        result = preview_preparation_source(workspace_id="ws-1", source_id=source_id, offset=0, limit=10, registry=registry)

        assert result.header_row_number is None
        assert result.column_labels == ["A", "B"]
        assert result.data_start_row is None
        assert all(r.in_active_region for r in result.rows)
        assert result.column_roles == ["not_assigned", "not_assigned"]
